from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, make_response, Blueprint
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timezone
from extensions import db
from models import Categoria, Producto, Proveedor, ProductoProveedor, Compra, DetalleCompra, Usuario, Carrito, Pedido, DetallePedido, Review, HistorialPedido
import re
from sqlalchemy import and_, exists, or_, func
from sqlalchemy.sql import expression
import pdfkit
from functools import wraps
from flask_cors import CORS
from sqlalchemy.exc import IntegrityError
from datetime import datetime
from sqlalchemy import and_
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO
import tempfile
import os
import pandas as pd  # Nueva importación para pandas
import xlsxwriter  # Nueva importación para xlsxwriter
import base64
from datetime import datetime, timezone, timedelta
from sqlalchemy import func, cast, Date
from collections import defaultdict
from flask import g
from zoneinfo import ZoneInfo
import random
from functools import wraps
from xhtml2pdf import pisa
from io import BytesIO

logging.basicConfig(level=logging.INFO)

app = Flask(__name__)
CORS(app)
app.secret_key = 'Cristian3012'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

def calcular_precio_final():
    return expression.case(
        (Producto.descuento > 0, Producto.precio * (1 - Producto.descuento / 100.0)),
        else_=Producto.precio)

def calcular_precio_final_producto(producto):
    if producto.descuento > 0:
        return producto.precio * (1 - producto.descuento / 100.0)
    return producto.precio

@app.context_processor
def inject_categorias_y_carrito():
    categorias_principales = Categoria.query.filter_by(parent_id=None, activa=True).all()
    carrito_count = None
    puntos_usuario = None
    forgot_errors = session.pop('forgot_errors', None)
    forgot_form_data = session.pop('forgot_form_data', None)

    if 'usuario_id' in session:
        carrito_count = Carrito.query.filter_by(usuario_id=session['usuario_id']).count()
        usuario = Usuario.query.get(session['usuario_id'])
        puntos_usuario = usuario.puntos if usuario else None
        
    return dict(
        categorias_principales=categorias_principales, 
        carrito_count=carrito_count,
        puntos_usuario=puntos_usuario,
        forgot_errors=forgot_errors, 
        forgot_form_data=forgot_form_data
    )

def eliminar_categoria_recursivamente(categoria):
    for subcat in categoria.subcategorias:
        eliminar_categoria_recursivamente(subcat)
    db.session.delete(categoria)

@app.before_request
def before_request():
    g.current_time = datetime.now(timezone.utc)


@app.route('/')
def index():
    categorias_principales = Categoria.query.filter_by(parent_id=None, activa=True).all()
    productos_destacados = Producto.query.filter_by(destacado=True, activo=True).join(Categoria).filter(Categoria.activa == True).filter(Producto.stock > 0).limit(8).all()
    return render_template('usuario/index.html', 
                          categorias=categorias_principales,
                          productos=productos_destacados)

@app.route('/categoria/<int:categoria_id>')
def mostrar_categoria(categoria_id):
    categoria = Categoria.query.filter_by(id=categoria_id, activa=True).first_or_404()
    subcategorias = Categoria.query.filter_by(parent_id=categoria_id, activa=True).all()
    
    if subcategorias:
        return render_template('usuario/categoria.html', 
                              categoria=categoria, 
                              subcategorias=subcategorias)
    else:
        min_precio = request.args.get('min_precio', type=float, default=0)
        max_precio = request.args.get('max_precio', type=float, default=10000)
        marca = request.args.get('marca', default=None)
        orden = request.args.get('orden', 'destacados')
        pagina = request.args.get('pagina', 1, type=int)
        por_pagina = 12
        
        if min_precio > max_precio:
            min_precio, max_precio = max_precio, min_precio
        
        query = Producto.query.filter_by(categoria_id=categoria_id, activo=True).join(Categoria).filter(Categoria.activa == True).filter(Producto.stock > 0)
        precio_final = calcular_precio_final()
        query = query.filter(precio_final >= min_precio, precio_final <= max_precio)
        
        if marca:
            query = query.filter(Producto.marca.ilike(f'%{marca}%'))
        
        if orden == 'precio_asc':
            query = query.order_by(precio_final.asc())
        elif orden == 'precio_desc':
            query = query.order_by(precio_final.desc())
        else:
            query = query.order_by(Producto.destacado.desc())
        
        paginacion = query.paginate(page=pagina, per_page=por_pagina, error_out=False)
        productos = paginacion.items
        
        marcas = db.session.query(Producto.marca).filter_by(categoria_id=categoria_id, activo=True).join(Categoria).filter(Categoria.activa == True).filter(Producto.stock > 0).distinct().all()
        marcas = [m[0] for m in marcas if m[0]]
        
        return render_template('usuario/productos.html', 
                              categoria=categoria, 
                              productos=productos,
                              paginacion=paginacion,
                              min_precio=min_precio,
                              max_precio=max_precio,
                              marca=marca,
                              marcas=marcas,
                              orden=orden)

@app.route('/buscar')
def buscar_productos():
    query = request.args.get('q', '').strip()
    
    productos = Producto.query.filter(
        Producto.nombre.ilike(f'{query}%'),
        Producto.activo == True
    ).join(Categoria).filter(Categoria.activa == True).limit(5).all()
    
    if productos:
        return redirect(url_for('todos_productos', q=query))
    else:
        return redirect(url_for('todos_productos', q=query, no_match=True))

@app.route('/todas_categorias')
def todas_categorias():
    categorias_principales = Categoria.query.filter_by(parent_id=None, activa=True).all()
    
    categorias_con_sub = []
    for cat in categorias_principales:
        subcategorias = Categoria.query.filter_by(parent_id=cat.id, activa=True).all()
        if subcategorias:
            categorias_con_sub.append({
                'categoria': cat,
                'subcategorias': subcategorias
            })
    
    return render_template('usuario/todas_categorias.html',
                          categorias_con_sub=categorias_con_sub,
                          categorias_principales=categorias_principales)

@app.route('/ofertas')
def todas_ofertas():
    min_precio = request.args.get('min_precio', type=float, default=0)
    max_precio = request.args.get('max_precio', type=float, default=10000)
    marca = request.args.get('marca', default=None)
    categoria_id = request.args.get('categoria_id', type=int, default=None)
    orden = request.args.get('orden', 'destacados')
    pagina = request.args.get('pagina', 1, type=int)
    por_pagina = 12
    
    if min_precio > max_precio:
        min_precio, max_precio = max_precio, min_precio
    
    query = Producto.query.filter(Producto.descuento > 0, Producto.activo == True, Producto.stock > 0).join(Categoria).filter(Categoria.activa == True)
    precio_final = calcular_precio_final()
    query = query.filter(precio_final >= min_precio, precio_final <= max_precio)
    
    if marca:
        query = query.filter(Producto.marca.ilike(f'%{marca}%'))
    
    if categoria_id:
        categoria = db.session.get(Categoria, categoria_id)
        if categoria and categoria.activa:
            subcategorias_ids = [c.id for c in Categoria.query.filter_by(parent_id=categoria_id, activa=True).all()]
            subcategorias_ids.append(categoria_id)
            query = query.filter(Producto.categoria_id.in_(subcategorias_ids))
    
    if orden == 'precio_asc':
        query = query.order_by(precio_final.asc())
    elif orden == 'precio_desc':
        query = query.order_by(precio_final.desc())
    else:
        query = query.order_by(Producto.descuento.desc())
    
    paginacion = query.paginate(page=pagina, per_page=por_pagina, error_out=False)
    productos = paginacion.items
    
    marcas = db.session.query(Producto.marca).filter(Producto.descuento > 0, Producto.activo == True).join(Categoria).filter(Categoria.activa == True).distinct().all()
    marcas = [m[0] for m in marcas if m[0]]
    
    categorias_principales = Categoria.query.filter_by(parent_id=None, activa=True).all()
    
    return render_template('usuario/ofertas.html', 
                          productos=productos,
                          paginacion=paginacion,
                          min_precio=min_precio,
                          max_precio=max_precio,
                          marca=marca,
                          marcas=marcas,
                          categorias_principales=categorias_principales,
                          categoria_id=categoria_id,
                          orden=orden)

@app.route('/destacados')
def todos_destacados():
    min_precio = request.args.get('min_precio', type=float, default=0)
    max_precio = request.args.get('max_precio', type=float, default=10000)
    marca = request.args.get('marca', default=None)
    categoria_id = request.args.get('categoria_id', type=int, default=None)
    orden = request.args.get('orden', 'destacados')
    pagina = request.args.get('pagina', 1, type=int)
    por_pagina = 12
    
    if min_precio > max_precio:
        min_precio, max_precio = max_precio, min_precio
    
    query = Producto.query.filter(Producto.destacado == True, Producto.activo == True, Producto.stock > 0).join(Categoria).filter(Categoria.activa == True)
    precio_final = calcular_precio_final()
    query = query.filter(precio_final >= min_precio, precio_final <= max_precio)
    
    if marca:
        query = query.filter(Producto.marca.ilike(f'%{marca}%'))
    
    if categoria_id:
        categoria = db.session.get(Categoria, categoria_id)
        if categoria and categoria.activa:
            subcategorias_ids = [c.id for c in Categoria.query.filter_by(parent_id=categoria_id, activa=True).all()]
            subcategorias_ids.append(categoria_id)
            query = query.filter(Producto.categoria_id.in_(subcategorias_ids))
    
    if orden == 'precio_asc':
        query = query.order_by(precio_final.asc())
    elif orden == 'precio_desc':
        query = query.order_by(precio_final.desc())
    else:
        query = query.order_by(Producto.destacado.desc(), Producto.nombre.asc())
    
    paginacion = query.paginate(page=pagina, per_page=por_pagina, error_out=False)
    productos = paginacion.items
    
    marcas = db.session.query(Producto.marca).filter(Producto.destacado == True, Producto.activo == True).join(Categoria).filter(Categoria.activa == True).distinct().all()
    marcas = [m[0] for m in marcas if m[0]]
    
    categorias_principales = Categoria.query.filter_by(parent_id=None, activa=True).all()
    
    for categoria in categorias_principales:
        categoria.tiene_destacados = db.session.query(
            exists().where(and_(
                Producto.categoria_id == categoria.id,
                Producto.destacado == True,
                Producto.activo == True,
                Categoria.activa == True
            ))
        ).scalar()
        
        for subcategoria in categoria.subcategorias:
            if subcategoria.activa:
                subcategoria.tiene_destacados = db.session.query(
                    exists().where(and_(
                        Producto.categoria_id == subcategoria.id,
                        Producto.destacado == True,
                        Producto.activo == True,
                        Categoria.activa == True
                    ))
                ).scalar()
    
    return render_template('usuario/destacados.html', 
                          productos=productos,
                          paginacion=paginacion,
                          min_precio=min_precio,
                          max_precio=max_precio,
                          marca=marca,
                          marcas=marcas,
                          categorias_principales=categorias_principales,
                          categoria_id=categoria_id,
                          orden=orden)

@app.route('/todos_productos')
def todos_productos():
    min_precio = request.args.get('min_precio', type=float, default=0)
    max_precio = request.args.get('max_precio', type=float, default=10000)
    marca = request.args.get('marca', default=None)
    categoria_id = request.args.get('categoria_id', type=int, default=None)
    orden = request.args.get('orden', 'destacados')
    pagina = request.args.get('pagina', 1, type=int)
    por_pagina = 12
    search_query = request.args.get('q', '').strip()
    no_match = request.args.get('no_match', False, type=bool)

    if min_precio > max_precio:
        min_precio, max_precio = max_precio, min_precio

    query = Producto.query.filter(Producto.activo == True, Producto.stock > 0).join(Categoria).filter(Categoria.activa == True)
    precio_final = calcular_precio_final()
    query = query.filter(precio_final >= min_precio, precio_final <= max_precio)

    if search_query:
        query = query.filter(Producto.nombre.ilike(f'{search_query}%'))

    if marca:
        query = query.filter(Producto.marca.ilike(f'%{marca}%'))

    if categoria_id:
        categoria = db.session.get(Categoria, categoria_id)
        if categoria and categoria.activa:
            subcategorias_ids = [c.id for c in Categoria.query.filter_by(parent_id=categoria_id, activa=True).all()]
            subcategorias_ids.append(categoria_id)
            query = query.filter(Producto.categoria_id.in_(subcategorias_ids))

    if orden == 'precio_asc':
        query = query.order_by(precio_final.asc())
    elif orden == 'precio_desc':
        query = query.order_by(precio_final.desc())
    else:
        query = query.order_by(Producto.destacado.desc(), Producto.nombre.asc())

    paginacion = query.paginate(page=pagina, per_page=por_pagina, error_out=False)
    productos = paginacion.items

    marcas = db.session.query(Producto.marca).filter(Producto.activo == True, Producto.stock > 0).join(Categoria).filter(Categoria.activa == True).distinct().all()
    marcas = [m[0] for m in marcas if m[0]]
    categorias_principales = Categoria.query.filter_by(parent_id=None, activa=True).all()

    return render_template('usuario/todos_productos.html', 
                          productos=productos,
                          paginacion=paginacion,
                          min_precio=min_precio,
                          max_precio=max_precio,
                          marca=marca,
                          marcas=marcas,
                          categorias_principales=categorias_principales,
                          categoria_id=categoria_id,
                          orden=orden,
                          search_query=search_query,
                          no_match=no_match)

@app.route('/sugerencias-busqueda')
def sugerencias_busqueda():
    query = request.args.get('q', '').strip()
    results = []
    
    if query:
        productos = Producto.query.filter(
            Producto.nombre.ilike(f'{query}%')
        ).limit(5).all()
        
        results = [{
            'id': p.id,
            'nombre': p.nombre
        } for p in productos]
    
    return jsonify(results)

@app.route('/producto/<int:producto_id>')
def detalle_producto(producto_id):
    producto = Producto.query.filter(Producto.id == producto_id, Producto.activo == True, Producto.stock > 0).join(Categoria).filter(Categoria.activa == True).first_or_404()
    
    reviews = Review.query.filter_by(producto_id=producto_id).join(Usuario).all()
    
    promedio_calificacion = 0
    if reviews:
        promedio_calificacion = sum(review.calificacion for review in reviews) / len(reviews)
    
    user_review_count = 0
    if 'usuario_id' in session:
        user_review_count = Review.query.filter_by(
            usuario_id=session['usuario_id'],
            producto_id=producto_id
        ).count()
    
    return render_template('usuario/detalle_producto.html', 
                          producto=producto,
                          reviews=reviews,
                          promedio_calificacion=promedio_calificacion,
                          total_reviews=len(reviews),
                          user_review_count=user_review_count)

@app.route('/agregar_carrito/<int:producto_id>', methods=['POST'])
def agregar_carrito(producto_id):
    if 'usuario_id' not in session:
        flash('Debes iniciar sesion para agregar productos al carrito', 'danger')
        return redirect(url_for('index') + '#loginModal')
    
    producto = Producto.query.filter_by(id=producto_id, activo=True).join(Categoria).filter(Categoria.activa == True).first_or_404()
    cantidad = int(request.form.get('cantidad', 1))
    
    if cantidad > producto.stock:
        flash(f'No hay suficiente stock. Solo hay {producto.stock} unidades disponibles.', 'danger')
        return redirect(url_for('detalle_producto', producto_id=producto_id))
    
    item = Carrito.query.filter_by(
        usuario_id=session['usuario_id'], 
        producto_id=producto_id
    ).first()
    
    if item:
        nueva_cantidad = item.cantidad + cantidad
        if nueva_cantidad > producto.stock:
            flash(f'No puedes agregar mas de {producto.stock} unidades de este producto.', 'danger')
            return redirect(url_for('detalle_producto', producto_id=producto_id))
        item.cantidad = nueva_cantidad
    else:
        nuevo_item = Carrito(
            usuario_id=session['usuario_id'],
            producto_id=producto_id,
            cantidad=cantidad,
            fecha_agregado=datetime.now(timezone.utc)
        )
        db.session.add(nuevo_item)
    
    db.session.commit()
    flash('Producto agregado al carrito', 'success')
    return redirect(url_for('detalle_producto', producto_id=producto_id))

@app.route('/carrito')
def ver_carrito():
    if 'usuario_id' not in session:
        flash('Debes iniciar sesion para ver el carrito', 'danger')
        return redirect(url_for('index') + '#loginModal')
    
    items = Carrito.query.filter_by(usuario_id=session['usuario_id']).all()
    
    subtotal_sin_descuento = 0
    total = 0
    descuento_total = 0

    for item in items:
        precio_original = item.producto.precio
        descuento = item.producto.descuento
        
        precio_final = precio_original * (1 - descuento / 100) if descuento > 0 else precio_original
        
        subtotal_sin_descuento += precio_original * item.cantidad
        total += precio_final * item.cantidad
        
        if descuento > 0:
            descuento_total += (precio_original - precio_final) * item.cantidad
    
    return render_template(
        'usuario/carrito.html',
        items=items,
        subtotal_sin_descuento=subtotal_sin_descuento,
        total=total,
        descuento_total=descuento_total
    )

@app.route('/eliminar_carrito/<int:item_id>')
def eliminar_carrito(item_id):
    if 'usuario_id' not in session:
        return jsonify({'success': False, 'message': 'Debes iniciar sesion para eliminar productos'}), 401
    
    item = Carrito.query.get_or_404(item_id)
    if item.usuario_id != session.get('usuario_id'):
        return jsonify({'success': False, 'message': 'No tienes permiso para realizar esta accion'}), 403

    db.session.delete(item)
    db.session.commit()
    
    items = Carrito.query.filter_by(usuario_id=session['usuario_id']).all()
    subtotal_sin_descuento = 0
    total = 0
    descuento_total = 0

    for item in items:
        precio_original = item.producto.precio
        descuento = item.producto.descuento
        precio_final = precio_original * (1 - descuento / 100) if descuento > 0 else precio_original
        
        subtotal_sin_descuento += precio_original * item.cantidad
        total += precio_final * item.cantidad
        if descuento > 0:
            descuento_total += (precio_original - precio_final) * item.cantidad

    return jsonify({
        'success': True,
        'message': 'Producto eliminado del carrito',
        'cart_totals': {
            'subtotal_sin_descuento': round(subtotal_sin_descuento, 2),
            'descuento_total': round(descuento_total, 2),
            'total': round(total, 2)
        }
    })

@app.route('/actualizar_carrito/<int:item_id>', methods=['POST'])
def actualizar_carrito(item_id):
    if 'usuario_id' not in session:
        return jsonify({'success': False, 'message': 'Debes iniciar sesion para modificar el carrito'}), 401
    
    item = Carrito.query.get_or_404(item_id)
    if item.usuario_id != session.get('usuario_id'):
        return jsonify({'success': False, 'message': 'No tienes permiso para realizar esta accion'}), 403
    
    cantidad = int(request.form.get('cantidad', 1))
    previous_quantity = item.cantidad
    
    if cantidad > item.producto.stock:
        return jsonify({
            'success': False,
            'message': f'No puedes tener mas de {item.producto.stock} unidades de este producto en el carrito.',
            'previous_quantity': previous_quantity
        })
    
    precio_original = item.producto.precio
    descuento = item.producto.descuento
    precio_final = precio_original * (1 - descuento / 100) if descuento > 0 else precio_original
    
    if cantidad < 1:
        db.session.delete(item)
        db.session.commit()
        action = 'delete'
    else:
        item.cantidad = cantidad
        db.session.commit()
        action = 'update'
    
    items = Carrito.query.filter_by(usuario_id=session['usuario_id']).all()
    subtotal_sin_descuento = 0
    total = 0
    descuento_total = 0

    for item in items:
        precio_original = item.producto.precio
        descuento = item.producto.descuento
        precio_final_item = precio_original * (1 - descuento / 100) if descuento > 0 else precio_original
        
        subtotal_sin_descuento += precio_original * item.cantidad
        total += precio_final_item * item.cantidad
        if descuento > 0:
            descuento_total += (precio_original - precio_final_item) * item.cantidad

    response_data = {
        'success': True,
        'action': action,
        'item_id': item_id,
        'cart_totals': {
            'subtotal_sin_descuento': round(subtotal_sin_descuento, 2),
            'descuento_total': round(descuento_total, 2),
            'total': round(total, 2)
        }
    }
    
    if action == 'update':
        response_data['new_quantity'] = cantidad
        response_data['new_subtotal'] = round(precio_final * cantidad, 2)
        
    return jsonify(response_data)

@app.route('/checkout', methods=['GET', 'POST'])
def checkout():
    if 'usuario_id' not in session:
        return redirect(url_for('index') + '#loginModal')
    
    items = Carrito.query.filter_by(usuario_id=session['usuario_id']).all()
    if not items:
        flash('Tu carrito esta vacio', 'warning')
        return redirect(url_for('index'))

    total = 0
    descuento_total = 0
    descuento_puntos = 0

    for item in items:
        precio_final = calcular_precio_final_producto(item.producto)
        total += precio_final * item.cantidad
        
        if item.producto.descuento > 0:
            descuento = item.producto.precio * (item.producto.descuento / 100.0)
            descuento_total += descuento * item.cantidad
    
    descuento_puntos_input = request.form.get('descuento_puntos', 0, type=float)
    if descuento_puntos_input > 0:
        total -= descuento_puntos_input
    
    usuario = Usuario.query.get(session['usuario_id'])

    if request.method == 'POST':
        nombre = request.form.get('nombre')
        direccion = request.form.get('direccion')
        ciudad = request.form.get('ciudad')
        codigo_postal = request.form.get('codigo_postal')
        telefono = request.form.get('telefono')
        metodo_pago = request.form.get('metodo_pago')

        if not all([nombre, direccion, ciudad, codigo_postal, telefono]):
            flash('Por favor, completa todos los campos de la direccion de envio.', 'danger')
            return redirect(url_for('checkout'))

        if metodo_pago not in ['tarjeta', 'contraentrega']:
            flash('Metodo de pago no valido.', 'danger')
            return redirect(url_for('checkout'))

        if metodo_pago == 'tarjeta':
            numero_tarjeta = request.form.get('numero_tarjeta')
            fecha_expiracion = request.form.get('fecha_expiracion')
            cvv = request.form.get('cvv')
            nombre_tarjeta = request.form.get('nombre_tarjeta')

            if not all([numero_tarjeta, fecha_expiracion, cvv, nombre_tarjeta]):
                flash('Por favor, completa todos los campos de la tarjeta.', 'danger')
                return redirect(url_for('checkout'))

            if not re.match(r'^\d{16}$', numero_tarjeta.replace(' ', '')):
                flash('Numero de tarjeta no valido.', 'danger')
                return redirect(url_for('checkout'))

            if not re.match(r'^\d{2}/\d{2}$', fecha_expiracion):
                flash('Fecha de expiracion no valida. Usa el formato MM/AA.', 'danger')
                return redirect(url_for('checkout'))

            try:
                mes, anio = map(int, fecha_expiracion.split('/'))
                anio = 2000 + anio
                now = datetime.now(timezone.utc)
                if anio < now.year or (anio == now.year and mes < now.month):
                    flash('La tarjeta ha expirado.', 'danger')
                    return redirect(url_for('checkout'))
                if mes < 1 or mes > 12:
                    flash('Mes de expiracion no valido.', 'danger')
                    return redirect(url_for('checkout'))
            except ValueError:
                flash('Fecha de expiracion no valida.', 'danger')
                return redirect(url_for('checkout'))

            if not re.match(r'^\d{3,4}$', cvv):
                flash('CVV no valido.', 'danger')
                return redirect(url_for('checkout'))

            if not nombre_tarjeta.strip():
                flash('El nombre en la tarjeta es requerido.', 'danger')
                return redirect(url_for('checkout'))

        for item in items:
            if item.cantidad > item.producto.stock:
                flash(f'No hay suficiente stock para {item.producto.nombre}. Disponible: {item.producto.stock}, Requerido: {item.cantidad}', 'danger')
                return redirect(url_for('checkout'))

        puntos_usados = int(descuento_puntos_input * 10)
        # Calcular puntos ganados (10 puntos por cada $100)
        puntos_ganados = int(total // 10)

        nuevo_pedido = Pedido(
            usuario_id=session['usuario_id'],
            fecha_pedido=datetime.now(timezone.utc),
            total=total,
            nombre=nombre,
            direccion=direccion,
            ciudad=ciudad,
            codigo_postal=codigo_postal,
            telefono=telefono,
            metodo_pago=metodo_pago,
            estado='pendiente',
            puntos_usados=puntos_usados,
            puntos_ganados=puntos_ganados  # Registrar puntos ganados
        )
        db.session.add(nuevo_pedido)
        db.session.flush()
        
        for item in items:
            precio_final = calcular_precio_final_producto(item.producto)
            detalle = DetallePedido(
                pedido_id=nuevo_pedido.id,
                producto_id=item.producto_id,
                cantidad=item.cantidad,
                precio=precio_final,
                descuento_aplicado=item.producto.descuento
            )
            db.session.add(detalle)
            db.session.delete(item)
        
        usuario = Usuario.query.get(session['usuario_id'])
        if puntos_usados > 0:
            if usuario.puntos >= puntos_usados:
                usuario.puntos -= puntos_usados
            else:
                db.session.rollback()
                flash('No tienes suficientes puntos para aplicar este descuento.', 'danger')
                return redirect(url_for('checkout'))

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f'Error al procesar el pedido: {str(e)}', 'danger')
            return redirect(url_for('checkout'))

        flash_message = f'Compra realizada con éxito. Tu pedido está en estado pendiente. Ganarás {puntos_ganados} puntos cuando se complete.'
        if puntos_usados > 0:
            flash_message += f' Usaste {puntos_usados} puntos.'
        flash(flash_message, 'success')
        
        return redirect(url_for('mis_pedidos'))
    
    return render_template('usuario/checkout.html', 
                         items=items, 
                         total=total, 
                         descuento_total=descuento_total, 
                         puntos_usuario=usuario.puntos)


@app.route('/canjear_puntos', methods=['POST'])
def canjear_puntos():
    if 'usuario_id' not in session:
        return jsonify({'success': False, 'message': 'Debes iniciar sesión'}), 401
    
    usuario = Usuario.query.get(session['usuario_id'])
    puntos_a_canjear = int(request.form.get('puntos', 0))
    
    if puntos_a_canjear <= 0:
        return jsonify({'success': False, 'message': 'Cantidad inválida'})
    
    if puntos_a_canjear > usuario.puntos:
        return jsonify({'success': False, 'message': 'No tienes suficientes puntos'})
    
    items = Carrito.query.filter_by(usuario_id=session['usuario_id']).all()
    total = 0
    for item in items:
        precio_final = calcular_precio_final_producto(item.producto)
        total += precio_final * item.cantidad
    
    descuento = puntos_a_canjear / 10
    if descuento > total:
        return jsonify({'success': False, 'message': 'El descuento por puntos no puede exceder el total del carrito'})
    
    return jsonify({
        'success': True,
        'descuento': descuento,
        'puntos_restantes': usuario.puntos - puntos_a_canjear
    })

@app.route('/actualizar_direccion/<int:pedido_id>', methods=['POST'])
def actualizar_direccion(pedido_id):
    if 'usuario_id' not in session:
        flash('Debes iniciar sesion para modificar la direccion.', 'danger')
        return redirect(url_for('index') + '#loginModal')
    
    pedido = Pedido.query.get_or_404(pedido_id)
    if pedido.usuario_id != session.get('usuario_id'):
        flash('No tienes permiso para modificar este pedido.', 'danger')
        return redirect(url_for('index'))
    
    nombre = request.form.get('nombre')
    direccion = request.form.get('direccion')
    ciudad = request.form.get('ciudad')
    codigo_postal = request.form.get('codigo_postal')
    telefono = request.form.get('telefono')

    if not all([nombre, direccion, ciudad, codigo_postal, telefono]):
        flash('Por favor, completa todos los campos de la direccion.', 'danger')
        return redirect(url_for('detalle_pedido', pedido_id=pedido_id))

    pedido.nombre = nombre
    pedido.direccion = direccion
    pedido.ciudad = ciudad
    pedido.codigo_postal = codigo_postal
    pedido.telefono = telefono

    db.session.commit()
    flash('Direccion actualizada con exito.', 'success')
    return redirect(url_for('detalle_pedido', pedido_id=pedido_id))

@app.route('/mis_pedidos')
def mis_pedidos():
    if 'usuario_id' not in session:
        flash('Debes iniciar sesion para ver tus pedidos', 'danger')
        return redirect(url_for('index') + '#loginModal')
    
    pedidos = Pedido.query.filter_by(usuario_id=session['usuario_id']).all()
    
    for pedido in pedidos:
        descuento_total = 0.0
        for detalle in pedido.detalles:
            if detalle.descuento_aplicado > 0:
                precio_original = detalle.precio / (1 - detalle.descuento_aplicado/100)
                descuento_unidad = precio_original - detalle.precio
                descuento_total += descuento_unidad * detalle.cantidad
        pedido.descuento_total = round(descuento_total, 2)
    
    return render_template('usuario/mis_pedidos.html', pedidos=pedidos)

@app.route('/pedido/<int:pedido_id>')
def detalle_pedido(pedido_id):
    if 'usuario_id' not in session:
        flash('Debes iniciar sesion para ver los detalles del pedido', 'danger')
        return redirect(url_for('index') + '#loginModal')
    
    pedido = Pedido.query.get_or_404(pedido_id)
    if pedido.usuario_id != session['usuario_id'] and not session.get('es_admin'):
        flash('No tienes permiso para ver este pedido', 'danger')
        return redirect(url_for('index'))
    
    detalles = DetallePedido.query.filter_by(pedido_id=pedido_id).all()
    
    subtotal_con_descuento = 0
    subtotal_sin_descuento = 0
    descuento_total = 0
    
    for detalle in detalles:
        precio_original = detalle.precio / (1 - detalle.descuento_aplicado/100) if detalle.descuento_aplicado > 0 else detalle.precio
        precio_final = detalle.precio
        
        subtotal_con_descuento += precio_final * detalle.cantidad
        subtotal_sin_descuento += precio_original * detalle.cantidad
        
        if detalle.descuento_aplicado > 0:
            descuento_total += (precio_original - precio_final) * detalle.cantidad
    
    if abs(pedido.total - subtotal_con_descuento) > 0.01:
        subtotal_con_descuento = pedido.total
    
    return render_template('usuario/detalle_pedido.html', 
                         pedido=pedido, 
                         detalles=detalles,
                         subtotal_con_descuento=subtotal_con_descuento,
                         subtotal_sin_descuento=subtotal_sin_descuento,
                         descuento_total=descuento_total)

@app.route('/registro', methods=['POST'])
def registro():
    errors = {}
    nombre = request.form.get('nombre')
    email = request.form.get('email')
    identificacion = request.form.get('identificacion')
    password = request.form.get('password')
    confirm_password = request.form.get('confirm_password')
    terms = request.form.get('terms')
    
    if not nombre:
        errors['nombre'] = 'Por favor ingresa tu nombre'
    if not email:
        errors['email'] = 'Por favor ingresa tu correo electrónico'
    if not identificacion:
        errors['identificacion'] = 'Por favor ingresa tu número de identificación'
    elif not re.match(r'^\d{6,12}$', identificacion):
        errors['identificacion'] = 'La identificación debe tener entre 6 y 12 dígitos numéricos'
    elif Usuario.query.filter_by(identificacion=identificacion).first():
        errors['identificacion'] = 'El número de identificación ya está registrado'
    if not password:
        errors['password'] = 'Por favor ingresa tu contraseña'
    if not confirm_password:
        errors['confirm_password'] = 'Por favor confirma tu contraseña'
    if not terms:
        errors['terms'] = 'Debes aceptar los Términos y Condiciones'
    
    if password and confirm_password and password != confirm_password:
        errors['confirm_password'] = 'Las contraseñas no coinciden'
    
    if email and Usuario.query.filter_by(email=email).first():
        errors['email'] = 'El correo electrónico ya está registrado'

    if len(password) < 6:
        errors['password'] = 'La contraseña debe tener al menos 6 caracteres'
    
    if errors:
        return jsonify({'success': False, 'errors': errors})
    
    nuevo_usuario = Usuario(
        nombre=nombre,
        email=email,
        identificacion=identificacion,
        password=generate_password_hash(password),
        fecha_registro=datetime.now(timezone.utc),
        puntos=100  # Asignar 100 puntos al nuevo usuario
    )
    db.session.add(nuevo_usuario)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': 'Registro exitoso. Se han asignado 100 puntos de bienvenida. Ahora puedes iniciar sesión'
    })

@app.route('/login', methods=['POST'])
def login():
    errors = {}
    email = request.form.get('email')
    password = request.form.get('password')
    
    if not email:
        errors['email'] = 'Por favor ingresa tu correo electronico'
    if not password:
        errors['password'] = 'Por favor ingresa tu contrasena'
    
    usuario = Usuario.query.filter_by(email=email, activo=True).first()
    if not usuario:
        errors['email'] = 'El correo electronico no esta registrado o la cuenta esta inactiva'
    if not usuario and not errors.get('email'):
        errors['email'] = 'El correo electronico no esta registrado'
    elif usuario and not check_password_hash(usuario.password, password) and not errors.get('password'):
        errors['password'] = 'Contrasena incorrecta'
    
    if errors:
        return jsonify({'success': False, 'errors': errors})
    
    session['usuario_id'] = usuario.id
    session['usuario_nombre'] = usuario.nombre
    session['es_admin'] = usuario.es_admin
    
    if usuario.es_admin:
        return jsonify({
            'success': True,
            'message': 'Inicio de sesion exitoso como administrador',
            'redirect': url_for('admin_inicio')
        })
    else:
        return jsonify({
            'success': True,
            'message': 'Inicio de sesion exitoso',
            'redirect': url_for('index')
        })

@app.route('/perfil')
def perfil():
    if 'usuario_id' not in session:
        flash('Debes iniciar sesion para ver tu perfil', 'danger')
        return redirect(url_for('index') + '#loginModal')
    
    usuario = Usuario.query.get_or_404(session['usuario_id'])
    return render_template('usuario/perfil.html', usuario=usuario)

@app.route('/actualizar_perfil', methods=['POST'])
def actualizar_perfil():
    if 'usuario_id' not in session:
        flash('Debes iniciar sesión para editar tu perfil', 'danger')
        return redirect(url_for('login'))
    
    usuario = Usuario.query.get_or_404(session['usuario_id'])
    
    nombre = request.form['nombre']
    email = request.form['email']
    identificacion = request.form['identificacion']
    
    if not nombre:
        flash('El nombre es obligatorio', 'danger')
        return redirect(url_for('perfil'))
    if not email:
        flash('El correo electrónico es obligatorio', 'danger')
        return redirect(url_for('perfil'))
    if not identificacion:
        flash('El número de identificación es obligatorio', 'danger')
        return redirect(url_for('perfil'))
    if not re.match(r'^\d{6,12}$', identificacion):
        flash('La identificación debe tener entre 6 y 12 dígitos numéricos', 'danger')
        return redirect(url_for('perfil'))
    if identificacion != usuario.identificacion and Usuario.query.filter_by(identificacion=identificacion).first():
        flash('El número de identificación ya está registrado', 'danger')
        return redirect(url_for('perfil'))
    if email != usuario.email and Usuario.query.filter_by(email=email).first():
        flash('El correo electrónico ya está registrado', 'danger')
        return redirect(url_for('perfil'))
    
    usuario.nombre = nombre
    usuario.email = email
    usuario.identificacion = identificacion
    
    password_actual = request.form.get('password_actual')
    nueva_password = request.form.get('nueva_password')
    confirmar_password = request.form.get('confirmar_password')
    
    if password_actual and nueva_password and confirmar_password:
        if not check_password_hash(usuario.password, password_actual):
            flash('La contraseña actual no es correcta', 'danger')
            return redirect(url_for('perfil'))
        
        if nueva_password != confirmar_password:
            flash('Las nuevas contraseñas no coinciden', 'danger')
            return redirect(url_for('perfil'))
        
        usuario.password = generate_password_hash(nueva_password)
    
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        flash('El número de identificación ya está registrado', 'danger')
        return redirect(url_for('perfil'))
    
    session['usuario_nombre'] = usuario.nombre
    
    flash('Perfil actualizado con éxito', 'success')
    return redirect(url_for('perfil'))

@app.route('/eliminar_cuenta', methods=['POST'])
def eliminar_cuenta():
    if 'usuario_id' not in session:
        flash('Debes iniciar sesion para realizar esta accion', 'danger')
        return redirect(url_for('login'))
    
    usuario = Usuario.query.get_or_404(session['usuario_id'])
    
    password = request.form.get('password')
    if not check_password_hash(usuario.password, password):
        flash('Contrasena incorrecta. No se pudo eliminar la cuenta.', 'danger')
        return redirect(url_for('perfil'))
    
    try:
        Carrito.query.filter_by(usuario_id=usuario.id).delete()
        Review.query.filter_by(usuario_id=usuario.id).delete()
        
        pedidos = Pedido.query.filter_by(usuario_id=usuario.id).all()
        for pedido in pedidos:
            DetallePedido.query.filter_by(pedido_id=pedido.id).delete()
            db.session.delete(pedido)
        
        db.session.delete(usuario)
        db.session.commit()
        
        session.clear()
        flash('Tu cuenta ha sido eliminada exitosamente', 'success')
        return redirect(url_for('index'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar la cuenta: {str(e)}', 'danger')
        return redirect(url_for('perfil'))

@app.route('/agregar_review/<int:producto_id>', methods=['POST'])
def agregar_review(producto_id):
    if 'usuario_id' not in session:
        flash('Debes iniciar sesion para agregar resenas', 'danger')
        return redirect(url_for('detalle_producto', producto_id=producto_id))

    review_count = db.session.query(Review).filter_by(
        usuario_id=session['usuario_id'],
        producto_id=producto_id
    ).count()
    
    if review_count >= 3:
        flash('Has alcanzado el limite de tres resenas por producto', 'warning')
        return redirect(url_for('detalle_producto', producto_id=producto_id))

    calificacion = int(request.form.get('calificacion'))
    comentario = request.form.get('comentario')
    
    nueva_review = Review(
        producto_id=producto_id,
        usuario_id=session['usuario_id'],
        calificacion=calificacion,
        comentario=comentario
    )
    
    db.session.add(nueva_review)
    db.session.commit()
    flash('Resena agregada con exito', 'success')
    return redirect(url_for('detalle_producto', producto_id=producto_id))

@app.route('/editar_review/<int:review_id>', methods=['POST'])
def editar_review(review_id):
    if 'usuario_id' not in session:
        flash('Debes iniciar sesion para editar resenas', 'danger')
        return redirect(url_for('login'))

    review = Review.query.get_or_404(review_id)
    
    if review.usuario_id != session['usuario_id'] and not session.get('es_admin'):
        flash('No tienes permiso para editar esta resena', 'danger')
        return redirect(url_for('detalle_producto', producto_id=review.producto_id))

    review.calificacion = int(request.form.get('calificacion'))
    review.comentario = request.form.get('comentario')
    db.session.commit()
    
    flash('Resena actualizada con exito', 'success')
    return redirect(url_for('detalle_producto', producto_id=review.producto_id))

@app.route('/eliminar_review/<int:review_id>', methods=['POST'])
def eliminar_review(review_id):
    if 'usuario_id' not in session:
        flash('Debes iniciar sesion para eliminar resenas', 'danger')
        return redirect(url_for('login'))

    review = Review.query.get_or_404(review_id)
    producto_id = review.producto_id
    
    if review.usuario_id != session['usuario_id'] and not session.get('es_admin'):
        flash('No tienes permiso para eliminar esta resena', 'danger')
        return redirect(url_for('detalle_producto', producto_id=producto_id))

    db.session.delete(review)
    db.session.commit()
    flash('Resena eliminada con exito', 'success')
    return redirect(url_for('detalle_producto', producto_id=producto_id))

@app.route('/logout')
def logout():
    session.clear()
    flash('Has cerrado sesion', 'info')
    return redirect(url_for('index'))

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session or not session.get('es_admin'):
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({
                    'success': False,
                    'message': 'Acceso denegado: se requiere autenticación de administrador'
                }), 403
            flash('Acceso denegado', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    total_productos = Producto.query.count()
    total_categorias = Categoria.query.count()
    total_usuarios = Usuario.query.count()
    total_ventas = db.session.query(func.sum(Pedido.total)).filter(Pedido.estado == 'completado').scalar() or 0
    pedidos_pendientes = Pedido.query.filter_by(estado='pendiente').count()
    productos_mas_vendidos = db.session.query(
        Producto.nombre, 
        func.sum(DetallePedido.cantidad).label('total_vendido')
    ).join(DetallePedido).group_by(Producto.id).order_by(func.sum(DetallePedido.cantidad).desc()).limit(5).all()
    pedidos_recientes = Pedido.query.order_by(Pedido.fecha_pedido.desc()).limit(5).all()
    current_year = datetime.now(timezone.utc).year
    monthly_sales = db.session.query(
        func.strftime('%m', Pedido.fecha_pedido).label('mes'),
        func.sum(Pedido.total).label('total')
    ).filter(
        func.strftime('%Y', Pedido.fecha_pedido) == str(current_year),
        Pedido.estado == 'completado'
    ).group_by(func.strftime('%m', Pedido.fecha_pedido)).order_by('mes').all()
    sales_data = [0] * 12
    for mes, total in monthly_sales:
        sales_data[int(mes) - 1] = float(total or 0)
    categorias_populares = db.session.query(
        Categoria.nombre,
        func.sum(DetallePedido.cantidad).label('total_vendido')
    ).join(Producto, Producto.categoria_id == Categoria.id
    ).join(DetallePedido, DetallePedido.producto_id == Producto.id
    ).group_by(Categoria.id
    ).order_by(func.sum(DetallePedido.cantidad).desc()
    ).limit(5).all()
    categorias_data = {
        'labels': [cat[0] for cat in categorias_populares],
        'data': [int(cat[1]) for cat in categorias_populares]
    }
    return render_template('admin/dashboard.html', 
                          total_productos=total_productos,
                          total_categorias=total_categorias,
                          total_usuarios=total_usuarios,
                          total_ventas=total_ventas,
                          pedidos_pendientes=pedidos_pendientes,
                          productos_mas_vendidos=productos_mas_vendidos,
                          pedidos_recientes=pedidos_recientes,
                          monthly_sales=sales_data,
                          categorias_populares=categorias_data)

@app.route('/forgot-password', methods=['POST'])
def forgot_password():
    form_data = {
        'nombre': request.form.get('nombre', '').strip(),
        'email': request.form.get('email', '').strip().lower(),
        'password': request.form.get('password', ''),
        'confirm_password': request.form.get('confirm_password', '')
    }
    errors = {}

    if not form_data['nombre']:
        errors['nombre'] = 'Por favor ingresa tu nombre completo'
    if not form_data['email']:
        errors['email'] = 'Por favor ingresa tu correo electronico'
    elif not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', form_data['email']):
        errors['email'] = 'Correo electronico no valido'
    if not form_data['password']:
        errors['password'] = 'Por favor ingresa tu nueva contrasena'
    elif len(form_data['password']) < 6:
        errors['password'] = 'La contrasena debe tener al menos 6 caracteres'
    if not form_data['confirm_password']:
        errors['confirm_password'] = 'Por favor confirma tu nueva contrasena'
    elif form_data['password'] != form_data['confirm_password']:
        errors['confirm_password'] = 'Las contrasenas no coinciden'

    usuario = None
    if not errors:
        usuario = Usuario.query.filter(
            func.lower(Usuario.email) == form_data['email'],
            func.lower(Usuario.nombre) == form_data['nombre'].lower()
        ).first()
        if not usuario:
            errors['general'] = 'No se encontro una cuenta con estos datos. Verifica tu nombre y correo.'

    if errors:
        session['forgot_errors'] = errors
        session['forgot_form_data'] = form_data
        return jsonify({'success': False, 'errors': errors})
    
    usuario.password = generate_password_hash(form_data['password'])
    db.session.commit()
    
    if 'forgot_errors' in session:
        session.pop('forgot_errors')
    if 'forgot_form_data' in session:
        session.pop('forgot_form_data')
    
    return jsonify({
        'success': True,
        'message': 'Tu contrasena ha sido restablecida exitosamente. Ahora puedes iniciar sesion con tu nueva contrasena.'
    })

@app.route('/admin/categorias')
@admin_required
def admin_categorias():
    search_query = request.args.get('search', '').strip()
    query = Categoria.query.filter_by(parent_id=None)
    
    if search_query:
        subcat_ids = [c.id for c in Categoria.query.filter(Categoria.nombre.ilike(f'%{search_query}%')).all()]
        query = query.filter(
            or_(
                Categoria.nombre.ilike(f'%{search_query}%'),
                Categoria.id.in_(
                    db.session.query(Categoria.parent_id).filter(Categoria.id.in_(subcat_ids)).distinct()
                )
            )
        )
    
    categorias_principales = query.order_by(Categoria.nombre.asc()).all()
    categorias_con_sub = []
    
    for cat in categorias_principales:
        subcategorias = Categoria.query.filter_by(parent_id=cat.id)
        if search_query and cat.nombre.lower().find(search_query.lower()) != -1:
            subcategorias = subcategorias.order_by(Categoria.nombre.asc()).all()
        else:
            subcategorias = subcategorias.filter(Categoria.nombre.ilike(f'%{search_query}%')).order_by(Categoria.nombre.asc()).all()
        if subcategorias or (search_query and cat.nombre.lower().find(search_query.lower()) != -1):
            categorias_con_sub.append({
                'categoria': cat,
                'subcategorias': subcategorias
            })
    
    return render_template('admin/categorias.html', categorias_con_sub=categorias_con_sub)

@app.route('/admin/categorias/agregar', methods=['GET', 'POST'])
def agregar_categoria():
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        tipo_categoria = request.form.get('tipo_categoria')
        parent_id = request.form.get('parent_id')

        if not nombre:
            flash('El nombre de la categoria es obligatorio', 'danger')
            return redirect(url_for('agregar_categoria'))

        if tipo_categoria == 'subcategoria':
            if not parent_id:
                flash('Debes seleccionar una categoria padre para una subcategoria', 'danger')
                return redirect(url_for('agregar_categoria'))
            parent_id = int(parent_id)
            parent_categoria = Categoria.query.get(parent_id)
            if not parent_categoria or parent_categoria.parent_id is not None:
                flash('La categoria padre seleccionada no es valida', 'danger')
                return redirect(url_for('agregar_categoria'))
        else:
            parent_id = None

        nueva_categoria = Categoria(nombre=nombre, parent_id=parent_id)
        db.session.add(nueva_categoria)
        db.session.commit()
        flash('Categoria agregada con exito', 'success')
        return redirect(url_for('admin_categorias'))

    categorias = Categoria.query.filter_by(parent_id=None).all()
    return render_template('admin/agregar_categoria.html', categorias=categorias)

@app.route('/admin/categorias/editar/<int:categoria_id>', methods=['GET', 'POST'])
def editar_categoria(categoria_id):
    categoria = Categoria.query.get_or_404(categoria_id)
    
    if request.method == 'POST':
        categoria.nombre = request.form['nombre']
        
        if categoria.parent_id is not None:
            parent_id = request.form.get('parent_id')
            if parent_id:
                parent_id = int(parent_id)
                if parent_id == categoria.id:
                    flash('Una categoria no puede ser su propia padre', 'danger')
                    return redirect(url_for('editar_categoria', categoria_id=categoria.id))
                
                parent_categoria = Categoria.query.get(parent_id)
                if parent_categoria and parent_categoria.parent_id is not None:
                    flash('Solo las categorias principales pueden ser seleccionadas como padre', 'danger')
                    return redirect(url_for('editar_categoria', categoria_id=categoria.id))
                
                categoria.parent_id = parent_id
        else:
            if request.form.get('parent_id'):
                flash('Las categorias principales no pueden tener una categoria padre', 'danger')
                return redirect(url_for('editar_categoria', categoria_id=categoria.id))
            categoria.parent_id = None

        db.session.commit()
        flash('Categoria actualizada con exito', 'success')
        return redirect(url_for('admin_categorias'))

    categorias_principales = Categoria.query.filter_by(parent_id=None).filter(Categoria.id != categoria_id).all()
    return render_template('admin/editar_categoria.html', categoria=categoria, categorias_principales=categorias_principales)

@app.route('/admin/categorias/eliminar/<int:categoria_id>')
def eliminar_categoria(categoria_id):
    categoria = Categoria.query.get_or_404(categoria_id)
    eliminar_categoria_recursivamente(categoria)
    db.session.commit()
    flash('Categoria y todas sus subcategorias eliminadas con exito', 'success')
    return redirect(url_for('admin_categorias'))

def eliminar_categoria_recursivamente(categoria):
    for subcat in categoria.subcategorias:
        eliminar_categoria_recursivamente(subcat)
    
    productos = Producto.query.filter_by(categoria_id=categoria.id).all()
    for producto in productos:
        Carrito.query.filter_by(producto_id=producto.id).delete()
    
    Producto.query.filter_by(categoria_id=categoria.id).delete()
    db.session.delete(categoria)

@app.route('/admin/productos')
@admin_required
def admin_productos():
    search_query = request.args.get('search', '').strip()
    subcategoria_id = request.args.get('subcategoria_id', type=int)
    marca = request.args.get('marca', '').strip()
    descripcion = request.args.get('descripcion', '').strip()
    
    query = Producto.query.join(Categoria).filter(Categoria.activa == True)
    
    if search_query:
        query = query.filter(Producto.nombre.ilike(f'{search_query}%'))
    if subcategoria_id:
        query = query.filter(Producto.categoria_id == subcategoria_id)
    if marca:
        query = query.filter(Producto.marca.ilike(f'%{marca}%'))
    if descripcion:
        query = query.filter(Producto.descripcion.ilike(f'%{descripcion}%'))
    
    productos = query.all()
    subcategorias = Categoria.query.filter(Categoria.parent_id != None, Categoria.activa == True).all()
    marcas = db.session.query(Producto.marca).filter(Producto.marca != None).distinct().all()
    marcas = [m[0] for m in marcas if m[0]]
    return render_template('admin/productos.html', productos=productos, subcategorias=subcategorias, marcas=marcas, marca=marca, descripcion=descripcion, search_query=search_query, subcategoria_id=subcategoria_id)


@app.route('/mis_compras')
def mis_compras():
    if 'usuario_id' not in session:
        flash('Debes iniciar sesión para ver tus compras', 'danger')
        return redirect(url_for('login'))
    
    # Obtener parámetros de filtrado
    fecha_inicio = request.args.get('fecha_inicio', '')
    fecha_fin = request.args.get('fecha_fin', '')
    metodo_pago = request.args.get('metodo_pago', '')
    
    query = Pedido.query.filter_by(usuario_id=session['usuario_id'], estado='completado')
    
    # Aplicar filtros
    try:
        if fecha_inicio:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            query = query.filter(Pedido.fecha_pedido >= fecha_inicio_dt)
        if fecha_fin:
            fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
            query = query.filter(Pedido.fecha_pedido <= fecha_fin_dt.replace(hour=23, minute=59, second=59))
    except ValueError:
        flash('Formato de fecha inválido. Use YYYY-MM-DD', 'danger')
        return redirect(url_for('mis_compras'))
    
    if metodo_pago:
        query = query.filter_by(metodo_pago=metodo_pago)
    
    # Ordenar por fecha descendente
    query = query.order_by(Pedido.fecha_pedido.desc())
    
    # Paginación
    pagina = request.args.get('pagina', 1, type=int)
    por_pagina = 10
    paginacion = query.paginate(page=pagina, per_page=por_pagina, error_out=False)
    compras = paginacion.items
    
    # Estadísticas
    total_compras = query.count()
    monto_total = query.with_entities(func.sum(Pedido.total)).scalar() or 0
    
    # Métodos de pago utilizados por el usuario
    metodos_pago = db.session.query(Pedido.metodo_pago).filter_by(
        usuario_id=session['usuario_id'], 
        estado='completado'
    ).distinct().all()
    metodos_pago = [m[0] for m in metodos_pago]
    
    return render_template('usuario/mis_compras.html',
                         compras=compras,
                         paginacion=paginacion,
                         total_compras=total_compras,
                         monto_total=monto_total,
                         metodos_pago=metodos_pago,
                         fecha_inicio=fecha_inicio,
                         fecha_fin=fecha_fin,
                         metodo_pago=metodo_pago)





@app.route('/admin/productos/agregar', methods=['GET', 'POST'])
@admin_required
def agregar_producto():
    def get_num_field(field, field_name, required=True, type_cast=float, default=0):
        value = request.form.get(field, None)
        if value is None or value.strip() == '':
            if required:
                flash(f'El campo "{field_name}" es obligatorio y no puede estar vacío.', 'danger')
                raise ValueError(f'Campo vacío: {field_name}')
            return default
        try:
            return type_cast(value)
        except Exception:
            flash(f'El campo "{field_name}" debe ser un número válido.', 'danger')
            raise ValueError(f'Campo inválido: {field_name}')

    if request.method == 'POST':
        try:
            # Obtener datos del formulario con validación robusta
            nombre = request.form.get('nombre', '').strip()
            descripcion = request.form.get('descripcion', '').strip()
            if not nombre or not descripcion:
                flash('Nombre y descripción son obligatorios', 'danger')
                return redirect(url_for('agregar_producto'))

            precio = get_num_field('precio', 'Precio', required=True, type_cast=float)
            costo = get_num_field('costo', 'Costo', required=False, type_cast=float, default=0)
            stock = get_num_field('stock', 'Stock', required=True, type_cast=int)
            stock_minimo = get_num_field('stock_minimo', 'Stock Mínimo', required=True, type_cast=int)
            stock_maximo = get_num_field('stock_maximo', 'Stock Máximo', required=True, type_cast=int)
            descuento = get_num_field('descuento', 'Descuento', required=False, type_cast=float, default=0)
            categoria_id = get_num_field('categoria_id', 'Subcategoría', required=True, type_cast=int)
            marca = request.form.get('marca', '').strip() or None
            imagen = request.form.get('imagen', 'https://via.placeholder.com/300').strip()
            destacado = 'destacado' in request.form
            proveedor_id = request.form.get('proveedor_id', None)
            producto_proveedor_id = request.form.get('producto_proveedor_id', None)

            # Validaciones lógicas
            if precio <= 0:
                flash('El precio debe ser mayor que cero', 'danger')
                return redirect(url_for('agregar_producto'))

            if stock < 0 or stock_minimo < 0 or stock_maximo < 0:
                flash('Stock, stock mínimo y stock máximo no pueden ser negativos', 'danger')
                return redirect(url_for('agregar_producto'))

            if stock_minimo > stock_maximo:
                flash('El stock mínimo no puede ser mayor que el stock máximo', 'danger')
                return redirect(url_for('agregar_producto'))

            if stock < stock_minimo or stock > stock_maximo:
                flash('El stock debe estar entre el mínimo y el máximo', 'danger')
                return redirect(url_for('agregar_producto'))

            if descuento < 0 or descuento > 100:
                flash('El descuento debe estar entre 0 y 100%', 'danger')
                return redirect(url_for('agregar_producto'))

            categoria = Categoria.query.get(categoria_id)
            if not categoria or not categoria.activa or categoria.parent_id is None:
                flash('Debe seleccionar una subcategoría válida', 'danger')
                return redirect(url_for('agregar_producto'))

            if proveedor_id:
                try:
                    proveedor_id = int(proveedor_id)
                except Exception:
                    flash('Proveedor no válido.', 'danger')
                    return redirect(url_for('agregar_producto'))
                proveedor = Proveedor.query.get(proveedor_id)
                if not proveedor or not proveedor.activo:
                    flash('Proveedor no válido o inactivo', 'danger')
                    return redirect(url_for('agregar_producto'))

                if costo and precio > 0:
                    margen = ((precio - costo) / costo) * 100 if costo else 0
                    if costo and margen < 50:
                        flash('El precio debe garantizar al menos un 50% de margen', 'danger')
                        return redirect(url_for('agregar_producto'))

                if producto_proveedor_id:
                    try:
                        producto_proveedor_id = int(producto_proveedor_id)
                    except Exception:
                        flash('Producto del proveedor no válido.', 'danger')
                        return redirect(url_for('agregar_producto'))
                    producto_proveedor = ProductoProveedor.query.filter_by(
                        id=producto_proveedor_id, proveedor_id=proveedor_id, activo=True
                    ).first()
                    if not producto_proveedor:
                        flash('Producto del proveedor no válido', 'danger')
                        return redirect(url_for('agregar_producto'))
                    if stock > producto_proveedor.stock:
                        flash(f'No hay suficiente stock en el proveedor (Disponible: {producto_proveedor.stock})', 'danger')
                        return redirect(url_for('agregar_producto'))

            # Crear nuevo producto
            nuevo_producto = Producto(
                nombre=nombre,
                descripcion=descripcion,
                precio=precio,
                costo=costo,
                stock=stock,
                stock_minimo=stock_minimo,
                stock_maximo=stock_maximo,
                descuento=descuento,
                categoria_id=categoria_id,
                marca=marca,
                imagen=imagen,
                destacado=destacado,
                proveedor_id=proveedor_id,
                fecha_creacion=datetime.now(timezone.utc),
                activo=True
            )

            db.session.add(nuevo_producto)

            # Actualizar stock del producto del proveedor si corresponde
            if producto_proveedor_id:
                producto_proveedor.stock -= stock
                db.session.add(producto_proveedor)

            db.session.commit()
            flash('Producto agregado con éxito', 'success')
            return redirect(url_for('admin_productos'))

        except ValueError as e:
            db.session.rollback()
            # El mensaje ya fue flasheado en la función auxiliar
        except IntegrityError as e:
            db.session.rollback()
            flash(f'Error al guardar el producto: {str(e)}', 'danger')
        except Exception as e:
            db.session.rollback()
            flash(f'Error inesperado: {str(e)}', 'danger')
        return redirect(url_for('agregar_producto'))

    # Método GET: Renderizar formulario
    subcategorias = Categoria.query.filter(Categoria.parent_id != None, Categoria.activa == True).all()
    proveedores = Proveedor.query.filter_by(activo=True).order_by(Proveedor.nombre).all()
    return render_template('admin/agregar_productos.html', subcategorias=subcategorias, proveedores=proveedores)





@app.route('/admin/inicio')
@admin_required
def admin_inicio():
    # Obtener nombre del administrador
    admin_nombre = session.get('usuario_nombre', 'Administrador')
    
    # Generar saludo dinámico según la hora
    hour = datetime.now(ZoneInfo('America/Bogota')).hour
    if hour < 12:
        saludo = "¡Buenos días"
    elif hour < 18:
        saludo = "¡Buenas tardes"
    else:
        saludo = "¡Buenas noches"
    
    # Consejo del día
    consejos = [
        "Revisa regularmente el inventario para mantener tus productos actualizados.",
        "Un mensaje de bienvenida cálido puede mejorar la experiencia de tus clientes.",
        "Analiza los comentarios de los usuarios para identificar oportunidades de mejora.",
        "Asegúrate de que las imágenes de los productos sean de alta calidad.",
        "Programa promociones especiales para fidelizar a tus clientes."
    ]
    consejo_dia = random.choice(consejos)
    
    # Log para depuración
    app.logger.info(f"Renderizando admin_inicio para usuario {admin_nombre}")
    
    return render_template('admin/inicio.html',
                          saludo=saludo,
                          admin_nombre=admin_nombre,
                          current_time=datetime.now(ZoneInfo('America/Bogota')),
                          consejo_dia=consejo_dia)









@app.route('/admin/compras/productos')
@admin_required
def admin_productos_comprados():
    search_query = request.args.get('search', '').strip()
    proveedor_id = request.args.get('proveedor_id', type=int)
    estado = request.args.get('estado', '')

    # Consulta base para productos comprados
    query = ProductoProveedor.query.filter_by(es_compra=True, activo=True).join(Proveedor).filter(
        Proveedor.activo == True
    )

    # Aplicar filtro de búsqueda
    if search_query:
        query = query.filter(ProductoProveedor.nombre.ilike(f'%{search_query}%'))

    # Aplicar filtro de proveedor
    if proveedor_id:
        query = query.filter(ProductoProveedor.proveedor_id == proveedor_id)

    # Aplicar filtro de estado
    if estado == 'disponible':
        query = query.filter(~exists().where(and_(
            DetalleCompra.producto_id == ProductoProveedor.id,
            DetalleCompra.convertido == True
        )))
    elif estado == 'convertido':
        query = query.filter(exists().where(and_(
            DetalleCompra.producto_id == ProductoProveedor.id,
            DetalleCompra.convertido == True
        )))

    # Obtener resultados ordenados
    productos_comprados = query.order_by(ProductoProveedor.fecha_creacion.desc()).all()
    proveedores = Proveedor.query.filter_by(activo=True).order_by(Proveedor.nombre).all()

    # Preprocesar productos para determinar si están convertidos
    for producto in productos_comprados:
        producto.is_convertido = any(d.convertido for d in producto.detalles_compra) if producto.detalles_compra else False

    # Depuración: Imprimir cantidad de productos encontrados
    print(f"Productos comprados encontrados: {len(productos_comprados)}")

    return render_template('admin/productos_comprados.html',
                         compras=productos_comprados,
                         proveedores=proveedores,
                         search_query=search_query,
                         proveedor_id=proveedor_id,
                         estado=estado)

@app.route('/admin/productos/eliminar/<int:producto_id>', methods=['POST'])
@admin_required
def eliminar_producto(producto_id):
    producto = Producto.query.get_or_404(producto_id)
    
    try:
        # Eliminar solo el producto de la tienda, sin afectar al proveedor
        db.session.delete(producto)
        db.session.commit()
        flash('Producto eliminado correctamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar el producto: {str(e)}', 'danger')
    
    return redirect(url_for('admin_productos'))


# Reemplazar la ruta /admin/productos/editar/<int:producto_id> en app.py con este bloque
@app.route('/admin/productos/editar/<int:producto_id>', methods=['GET', 'POST'])
@admin_required
def editar_producto(producto_id):
    producto = Producto.query.get_or_404(producto_id)
    subcategorias = Categoria.query.filter(Categoria.parent_id != None, Categoria.activa == True).all()
    
    costo_proveedor = None
    if producto.proveedor_id:
        producto_proveedor = ProductoProveedor.query.filter_by(
            nombre=producto.nombre, 
            proveedor_id=producto.proveedor_id
        ).first()
        if producto_proveedor:
            costo_proveedor = producto_proveedor.costo
    
    def get_num_field(field, field_name, required=True, type_cast=float, default=0):
        value = request.form.get(field, None)
        if value is None or value.strip() == '':
            if required:
                flash(f'El campo "{field_name}" es obligatorio y no puede estar vacío.', 'danger')
                raise ValueError(f'Campo vacío: {field_name}')
            return default
        try:
            return type_cast(value)
        except Exception:
            flash(f'El campo "{field_name}" debe ser un número válido.', 'danger')
            raise ValueError(f'Campo inválido: {field_name}')

    if request.method == 'POST':
        try:
            producto.nombre = request.form.get('nombre', '').strip()
            producto.descripcion = request.form.get('descripcion', '').strip()
            producto.marca = request.form.get('marca', '').strip() or None
            producto.precio = get_num_field('precio', 'Precio', required=True, type_cast=float)
            if not producto.proveedor_id:
                producto.categoria_id = get_num_field('categoria_id', 'Subcategoría', required=True, type_cast=int)
                producto.costo = get_num_field('costo', 'Costo', required=True, type_cast=float)
                if producto.costo <= 0:
                    flash('El costo debe ser mayor que cero', 'danger')
                    return redirect(url_for('editar_producto', producto_id=producto.id))
                margen = ((producto.precio - producto.costo) / producto.costo) * 100 if producto.costo else 0
                if producto.costo and margen < 50:
                    flash('El precio debe generar al menos un margen del 50% sobre el costo', 'danger')
                    return redirect(url_for('editar_producto', producto_id=producto.id))
            producto.stock = get_num_field('stock', 'Stock', required=True, type_cast=int)
            producto.descuento = get_num_field('descuento', 'Descuento', required=False, type_cast=float, default=0)
            producto.imagen = request.form.get('imagen', producto.imagen)
            producto.destacado = 'destacado' in request.form
            if not producto.proveedor_id:
                producto.stock_minimo = get_num_field('stock_minimo', 'Stock Mínimo', required=True, type_cast=int)
                producto.stock_maximo = get_num_field('stock_maximo', 'Stock Máximo', required=True, type_cast=int)
                if producto.stock_minimo > producto.stock_maximo:
                    flash('El stock mínimo no puede ser mayor que el stock máximo', 'danger')
                    return redirect(url_for('editar_producto', producto_id=producto.id))
                if producto.stock < producto.stock_minimo or producto.stock > producto.stock_maximo:
                    flash('El stock debe estar entre el mínimo y el máximo', 'danger')
                    return redirect(url_for('editar_producto', producto_id=producto.id))
            db.session.commit()
            flash('Producto actualizado con éxito', 'success')
            return redirect(url_for('admin_productos'))
        except ValueError:
            db.session.rollback()
            # El mensaje ya fue flasheado en la función auxiliar
        except IntegrityError:
            db.session.rollback()
            flash('Error al actualizar el producto', 'danger')
        return redirect(url_for('editar_producto', producto_id=producto.id))
    
    return render_template('admin/editar_productos.html', 
                         producto=producto, 
                         subcategorias=subcategorias, 
                         costo_proveedor=costo_proveedor)

@app.route('/descargar/factura/<int:pedido_id>')
def descargar_factura(pedido_id):
    # Verificar autenticación
    if 'usuario_id' not in session:
        flash('Debes iniciar sesión para descargar la factura', 'danger')
        return redirect(url_for('login'))
    
    # Obtener el pedido
    pedido = Pedido.query.get_or_404(pedido_id)
    
    # Verificar permisos (usuario dueño del pedido o admin)
    if pedido.usuario_id != session['usuario_id'] and not session.get('es_admin'):
        flash('No tienes permiso para descargar esta factura', 'danger')
        return redirect(url_for('mis_pedidos' if not session.get('es_admin') else 'admin_pedidos'))
    
    # Generar el HTML de la factura
    detalles = DetallePedido.query.filter_by(pedido_id=pedido_id).all()
    
    subtotal_con_descuento = 0
    subtotal_sin_descuento = 0
    descuento_total = 0
    
    for detalle in detalles:
        precio_original = detalle.precio / (1 - detalle.descuento_aplicado/100) if detalle.descuento_aplicado > 0 else detalle.precio
        precio_final = detalle.precio
        
        subtotal_con_descuento += precio_final * detalle.cantidad
        subtotal_sin_descuento += precio_original * detalle.cantidad
        
        if detalle.descuento_aplicado > 0:
            descuento_total += (precio_original - precio_final) * detalle.cantidad
    
    if abs(pedido.total - subtotal_con_descuento) > 0.01:
        subtotal_con_descuento = pedido.total
    
    html = render_template('usuario/factura.html', 
                         pedido=pedido, 
                         detalles=detalles,
                         subtotal_con_descuento=subtotal_con_descuento,
                         subtotal_sin_descuento=subtotal_sin_descuento,
                         descuento_total=descuento_total)
    
    try:
        # Configuración de pdfkit (asegúrate de tener wkhtmltopdf instalado)
        config = pdfkit.configuration(wkhtmltopdf=r'C:/Program Files/wkhtmltopdf/bin/wkhtmltopdf.exe')
        options = {
            'encoding': 'UTF-8',
            'quiet': '',
            'enable-local-file-access': '',
            'no-outline': None,
            'margin-top': '0mm',
            'margin-right': '0mm',
            'margin-bottom': '0mm',
            'margin-left': '0mm'
        }
        
        # Generar PDF
        pdf = pdfkit.from_string(html, False, configuration=config, options=options)
        
        # Crear respuesta con Content-Disposition: attachment para forzar descarga
        response = make_response(pdf)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=factura_{pedido.id}.pdf'
        
        return response
        
    except Exception as e:
        app.logger.error(f'Error al generar PDF: {str(e)}')
        flash(f'Error al generar el PDF: {str(e)}', 'danger')
        return redirect(url_for('detalle_pedido', pedido_id=pedido_id))


@app.route('/admin/usuarios')
@admin_required
def admin_usuarios():
    search_query = request.args.get('search', '').strip()
    rol = request.args.get('rol', '').strip()
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')

    query = Usuario.query

    if search_query:
        query = query.filter(
            or_(
                Usuario.nombre.ilike(f'{search_query}%'),
                Usuario.email.ilike(f'{search_query}%'),
                Usuario.identificacion.ilike(f'{search_query}%')
            )
        )

    if rol in ['admin', 'cliente']:
        query = query.filter(Usuario.es_admin == (rol == 'admin'))

    try:
        if fecha_desde:
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
            query = query.filter(Usuario.fecha_registro >= fecha_desde_dt)
        if fecha_hasta:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            query = query.filter(Usuario.fecha_registro <= fecha_hasta_dt.replace(hour=23, minute=59, second=59))
    except ValueError:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'Formato de fecha inválido'}), 400
        flash('Formato de fecha inválido', 'danger')
        return redirect(url_for('admin_usuarios'))

    usuarios = query.order_by(Usuario.fecha_registro.desc()).all()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        usuarios_data = []
        for usuario in usuarios:
            usuarios_data.append({
                'id': usuario.id,
                'nombre': usuario.nombre,
                'email': usuario.email,
                'identificacion': usuario.identificacion,
                'fecha_registro': usuario.fecha_registro.strftime('%d/%m/%Y'),
                'es_admin': usuario.es_admin,
                'activo': usuario.activo
            })
        return jsonify({'usuarios': usuarios_data})

    return render_template('admin/usuarios.html', usuarios=usuarios)

@app.route('/admin/usuarios/agregar', methods=['GET', 'POST'])
def agregar_usuario():
    if request.method == 'POST':
        errors = {}
        form_data = {
            'nombre': request.form.get('nombre', '').strip(),
            'email': request.form.get('email', '').strip().lower(),
            'identificacion': request.form.get('identificacion', '').strip(),
            'password': request.form.get('password', ''),
            'confirm_password': request.form.get('confirm_password', ''),
            'es_admin': request.form.get('es_admin', '0')
        }

        if not form_data['nombre']:
            errors['nombre'] = 'El nombre es obligatorio'
        elif len(form_data['nombre']) < 2:
            errors['nombre'] = 'El nombre debe tener al menos 2 caracteres'

        if not form_data['email']:
            errors['email'] = 'El correo electrónico es obligatorio'
        elif not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', form_data['email']):
            errors['email'] = 'Correo electrónico no válido'
        elif Usuario.query.filter_by(email=form_data['email']).first():
            errors['email'] = 'El correo electrónico ya está registrado'

        if not form_data['identificacion']:
            errors['identificacion'] = 'El número de identificación es obligatorio'
        elif not re.match(r'^\d{6,12}$', form_data['identificacion']):
            errors['identificacion'] = 'La identificación debe tener entre 6 y 12 dígitos numéricos'
        elif Usuario.query.filter_by(identificacion=form_data['identificacion']).first():
            errors['identificacion'] = 'El número de identificación ya está registrado'

        if not form_data['password']:
            errors['password'] = 'La contraseña es obligatoria'
        elif len(form_data['password']) < 6:
            errors['password'] = 'La contraseña debe tener al menos 6 caracteres'

        if not form_data['confirm_password']:
            errors['confirm_password'] = 'Debes confirmar la contraseña'
        elif form_data['password'] != form_data['confirm_password']:
            errors['confirm_password'] = 'Las contraseñas no coinciden'

        if errors:
            return render_template('admin/agregar_usuario.html', errors=errors, form_data=form_data)

        nuevo_usuario = Usuario(
            nombre=form_data['nombre'],
            email=form_data['email'],
            identificacion=form_data['identificacion'],
            password=generate_password_hash(form_data['password']),
            es_admin=form_data['es_admin'] == '1',
            fecha_registro=datetime.now(timezone.utc)
        )
        try:
            db.session.add(nuevo_usuario)
            db.session.commit()
            flash('Usuario agregado con éxito', 'success')
            return redirect(url_for('admin_usuarios'))
        except IntegrityError:
            db.session.rollback()
            errors['identificacion'] = 'El número de identificación ya está registrado'
            return render_template('admin/agregar_usuario.html', errors=errors, form_data=form_data)

    return render_template('admin/agregar_usuario.html', errors=None, form_data=None)

@app.route('/admin/usuarios/editar/<int:usuario_id>', methods=['GET', 'POST'])
def editar_usuario(usuario_id):
    usuario = Usuario.query.get_or_404(usuario_id)
    
    if request.method == 'POST':
        errors = {}
        form_data = {
            'nombre': request.form.get('nombre', '').strip(),
            'email': request.form.get('email', '').strip().lower(),
            'identificacion': request.form.get('identificacion', '').strip(),
            'password': request.form.get('password', ''),
            'confirm_password': request.form.get('confirm_password', ''),
            'es_admin': request.form.get('es_admin', '0')
        }

        if not form_data['nombre']:
            errors['nombre'] = 'El nombre es obligatorio'
        elif len(form_data['nombre']) < 2:
            errors['nombre'] = 'El nombre debe tener al menos 2 caracteres'

        if not form_data['email']:
            errors['email'] = 'El correo electrónico es obligatorio'
        elif not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', form_data['email']):
            errors['email'] = 'Correo electrónico no válido'
        elif form_data['email'] != usuario.email and Usuario.query.filter_by(email=form_data['email']).first():
            errors['email'] = 'El correo electrónico ya está registrado'

        if not form_data['identificacion']:
            errors['identificacion'] = 'El número de identificación es obligatorio'
        elif not re.match(r'^\d{6,12}$', form_data['identificacion']):
            errors['identificacion'] = 'La identificación debe tener entre 6 y 12 dígitos numéricos'
        elif form_data['identificacion'] != usuario.identificacion and Usuario.query.filter_by(identificacion=form_data['identificacion']).first():
            errors['identificacion'] = 'El número de identificación ya está registrado'

        if form_data['password'] or form_data['confirm_password']:
            if len(form_data['password']) < 6:
                errors['password'] = 'La contraseña debe tener al menos 6 caracteres'
            if form_data['password'] != form_data['confirm_password']:
                errors['confirm_password'] = 'Las contraseñas no coinciden'

        if errors:
            return render_template('admin/editar_usuario.html', usuario=usuario, errors=errors, form_data=form_data)

        usuario.nombre = form_data['nombre']
        usuario.email = form_data['email']
        usuario.identificacion = form_data['identificacion']
        usuario.es_admin = form_data['es_admin'] == '1'

        if form_data['password']:
            usuario.password = generate_password_hash(form_data['password'])

        try:
            db.session.commit()
            flash('Usuario actualizado con éxito', 'success')
            return redirect(url_for('admin_usuarios'))
        except IntegrityError:
            db.session.rollback()
            errors['identificacion'] = 'El número de identificación ya está registrado'
            return render_template('admin/editar_usuario.html', usuario=usuario, errors=errors, form_data=form_data)

    return render_template('admin/editar_usuario.html', usuario=usuario, errors=None, form_data=None)

@app.route('/admin/usuarios/eliminar/<int:usuario_id>')
def eliminar_usuario(usuario_id):
    usuario = Usuario.query.get_or_404(usuario_id)
    db.session.delete(usuario)
    db.session.commit()
    flash('Usuario eliminado con exito', 'success')
    return redirect(url_for('admin_usuarios'))

@app.route('/admin/pedidos')
@admin_required
def admin_pedidos():
    estado = request.args.get('estado', '').lower()
    fecha_inicio = request.args.get('fecha_inicio', '')
    fecha_fin = request.args.get('fecha_fin', '')
    valid_states = ['pendiente', 'completado', 'cancelado']

    query = Pedido.query

    if estado in valid_states:
        query = query.filter_by(estado=estado)

    try:
        if fecha_inicio:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            query = query.filter(Pedido.fecha_pedido >= fecha_inicio_dt)
        if fecha_fin:
            fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
            query = query.filter(Pedido.fecha_pedido <= fecha_fin_dt.replace(hour=23, minute=59, second=59))
    except ValueError:
        return jsonify({'success': False, 'message': 'Formato de fecha inválido'}), 400

    pedidos = query.all()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        pedidos_json = [{
            'id': pedido.id,
            'cliente': pedido.usuario.nombre,
            'fecha': pedido.fecha_pedido.strftime('%d %b, %Y'),
            'total': pedido.total,
            'metodo_pago': pedido.metodo_pago,
            'estado': pedido.estado
        } for pedido in pedidos]
        return jsonify({'pedidos': pedidos_json})

    return render_template('admin/pedidos.html', pedidos=pedidos)

@app.route('/admin/pedido/<int:pedido_id>')
def admin_detalle_pedido(pedido_id):
    pedido = Pedido.query.get_or_404(pedido_id)
    detalles = DetallePedido.query.filter_by(pedido_id=pedido_id).all()

    subtotal_con_descuento = 0
    subtotal_sin_descuento = 0
    descuento_total = 0
    
    for detalle in detalles:
        precio_original = detalle.precio / (1 - detalle.descuento_aplicado/100) if detalle.descuento_aplicado > 0 else detalle.precio
        precio_final = detalle.precio
        
        subtotal_con_descuento += precio_final * detalle.cantidad
        subtotal_sin_descuento += precio_original * detalle.cantidad
        
        if detalle.descuento_aplicado > 0:
            descuento_total += (precio_original - precio_final) * detalle.cantidad
    
    if abs(pedido.total - subtotal_con_descuento) > 0.01:
        subtotal_con_descuento = pedido.total
    
    return render_template('admin/detalle_pedido.html', 
                         pedido=pedido, 
                         detalles=detalles,
                         subtotal_con_descuento=subtotal_con_descuento,
                         subtotal_sin_descuento=subtotal_sin_descuento,
                         descuento_total=descuento_total)

@app.route('/admin/sales_data')
@admin_required
def sales_data():
    period = request.args.get('period', 'year')
    current_year = datetime.now(timezone.utc).year

    if period == 'month':
        query = db.session.query(
            func.strftime('%m', Pedido.fecha_pedido).label('mes'),
            func.sum(Pedido.total).label('total')
        ).filter(
            func.strftime('%Y', Pedido.fecha_pedido) == str(current_year),
            Pedido.estado == 'completado'
        ).group_by(func.strftime('%m', Pedido.fecha_pedido)).order_by('mes').all()

        labels = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        data = [0] * 12
        for mes, total in query:
            data[int(mes) - 1] = float(total or 0)

    elif period == 'quarter':
        query = db.session.query(
            func.strftime('%m', Pedido.fecha_pedido).label('mes'),
            func.sum(Pedido.total).label('total')
        ).filter(
            func.strftime('%Y', Pedido.fecha_pedido) == str(current_year),
            Pedido.estado == 'completado'
        ).group_by(func.strftime('%m', Pedido.fecha_pedido)).all()

        labels = ['Q1', 'Q2', 'Q3', 'Q4']
        data = [0] * 4
        for mes, total in query:
            month = int(mes)
            quarter = (month - 1) // 3
            data[quarter] += float(total or 0)

    else:
        query = db.session.query(
            func.strftime('%Y', Pedido.fecha_pedido).label('year'),
            func.sum(Pedido.total).label('total')
        ).filter(
            Pedido.estado == 'completado'
        ).group_by(func.strftime('%Y', Pedido.fecha_pedido)).order_by('year').all()

        start_year = current_year - 4
        labels = [str(year) for year in range(start_year, current_year + 1)]
        data = [0] * 5
        for year, total in query:
            if int(year) >= start_year:
                index = int(year) - start_year
                data[index] = float(total or 0)

    return jsonify({
        'labels': labels,
        'data': data
    })

@app.route('/admin/productos/toggle_activo/<int:producto_id>', methods=['POST'])
@admin_required
def toggle_activo_producto(producto_id):
    try:
        producto = Producto.query.get_or_404(producto_id)
        data = request.get_json()
        nuevo_estado = data.get('activo', False)
        
        producto.activo = nuevo_estado
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Estado del producto actualizado a {"activo" if nuevo_estado else "inactivo"}'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al actualizar el estado: {str(e)}'
        }), 500

@app.route('/admin/pedidos/cambiar_estado/<int:pedido_id>', methods=['POST'])
@admin_required
def cambiar_estado_pedido(pedido_id):
    try:
        pedido = Pedido.query.get_or_404(pedido_id)
        nuevo_estado = request.form.get('estado')

        if nuevo_estado not in ['pendiente', 'completado', 'cancelado']:
            return jsonify({'success': False, 'message': 'Estado no válido'}), 400

        if pedido.estado == nuevo_estado:
            return jsonify({'success': False, 'message': 'El pedido ya está en este estado'}), 400

        detalles = DetallePedido.query.filter_by(pedido_id=pedido_id).all()
        usuario = db.session.get(Usuario, pedido.usuario_id)
        estado_anterior = pedido.estado

        # Manejo de puntos y stock según transición de estados
        if nuevo_estado == 'completado':
            # Si pasamos a completado desde pendiente o cancelado
            if estado_anterior in ['pendiente', 'cancelado']:
                # Verificar stock y actualizar
                for detalle in detalles:
                    producto = db.session.get(Producto, detalle.producto_id)
                    if producto.stock < detalle.cantidad:
                        return jsonify({
                            'success': False,
                            'message': f'No hay suficiente stock para {producto.nombre}. Disponible: {producto.stock}, Requerido: {detalle.cantidad}'
                        }), 400
                    producto.stock -= detalle.cantidad
                
                # Asignar puntos (10 por cada $100)
                puntos_ganados = int(pedido.total // 10)
                if usuario:
                    usuario.puntos += puntos_ganados
                pedido.puntos_ganados = puntos_ganados

        elif nuevo_estado == 'cancelado':
            # Si cancelamos desde completado o pendiente
            if estado_anterior == 'completado':
                # Devolver stock y quitar puntos ganados
                for detalle in detalles:
                    producto = db.session.get(Producto, detalle.producto_id)
                    producto.stock = min(producto.stock + detalle.cantidad, producto.stock_maximo)
                
                if usuario:
                    usuario.puntos = max(usuario.puntos - pedido.puntos_ganados, 0)
                    if pedido.puntos_usados > 0:
                        usuario.puntos += pedido.puntos_usados
            
            elif estado_anterior == 'pendiente':
                # Solo devolver puntos usados si los había
                if usuario and pedido.puntos_usados > 0:
                    usuario.puntos += pedido.puntos_usados

        elif nuevo_estado == 'pendiente':
            # Si volvemos a pendiente desde completado
            if estado_anterior == 'completado':
                # Devolver stock y quitar puntos ganados
                for detalle in detalles:
                    producto = db.session.get(Producto, detalle.producto_id)
                    producto.stock = min(producto.stock + detalle.cantidad, producto.stock_maximo)
                
                if usuario:
                    usuario.puntos = max(usuario.puntos - pedido.puntos_ganados, 0)
                    if pedido.puntos_usados > 0:
                        usuario.puntos += pedido.puntos_usados

        # Actualizar estado del pedido
        pedido.estado = nuevo_estado
        
        # Registrar en el historial
        historial = HistorialPedido(
            pedido_id=pedido.id,
            estado_anterior=estado_anterior,
            estado_nuevo=nuevo_estado,
            usuario_id=session['usuario_id'],
            fecha_cambio=datetime.now(timezone.utc)
        )
        db.session.add(historial)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Estado del pedido #{pedido_id} cambiado a {nuevo_estado.capitalize()}',
            'estado': nuevo_estado,
            'puntos_ganados': pedido.puntos_ganados if nuevo_estado == 'completado' else 0,
            'puntos_usados': pedido.puntos_usados
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error inesperado: {str(e)}'}), 500

@app.route('/admin/categorias/toggle_activo/<int:categoria_id>', methods=['POST'])
@admin_required
def toggle_activo_categoria(categoria_id):
    try:
        categoria = Categoria.query.get_or_404(categoria_id)
        data = request.get_json()
        if data is None or 'activa' not in data:
            return jsonify({
                'success': False,
                'message': 'Datos inválidos: se requiere el campo "activa"'
            }), 400

        nuevo_estado = data['activa']

        def actualizar_estado_recursivo(cat, estado):
            cat.activa = estado
            if cat.parent_id is not None:
                Producto.query.filter_by(categoria_id=cat.id).update({'activo': estado})
            for subcat in cat.subcategorias:
                actualizar_estado_recursivo(subcat, estado)

        if not nuevo_estado and categoria.parent_id is None:
            subcategorias_activas = Categoria.query.filter_by(parent_id=categoria_id, activa=True).count()
            if subcategorias_activas > 0:
                return jsonify({
                    'success': False,
                    'message': 'No puedes desactivar una categoría principal con subcategorías activas'
                }), 400

        actualizar_estado_recursivo(categoria, nuevo_estado)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Estado de la categoría actualizado a {"activa" if nuevo_estado else "inactiva"}'
        })
        
    except IntegrityError as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error de integridad en la base de datos: {str(e)}'
        }), 500
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al actualizar el estado: {str(e)}'
        }), 500

@app.route('/admin/usuarios/toggle_activo/<int:usuario_id>', methods=['POST'])
@admin_required
def toggle_activo_usuario(usuario_id):
    try:
        usuario = Usuario.query.get_or_404(usuario_id)
        data = request.get_json()
        nuevo_estado = data.get('activo', False)
        
        if usuario.id == session.get('usuario_id') and not nuevo_estado:
            return jsonify({
                'success': False,
                'message': 'No puedes desactivar tu propia cuenta'
            }), 400
        
        usuario.activo = nuevo_estado
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Estado del usuario actualizado a {"activo" if nuevo_estado else "inactivo"}'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al actualizar el estado: {str(e)}'
        }), 500

@app.route('/admin/notificaciones')
@admin_required
def admin_notificaciones():
    from datetime import timedelta
    
    productos_bajo_stock = Producto.query.filter(
        Producto.stock < Producto.stock_minimo,
        Producto.activo == True
    ).all()
    
    pedidos_pendientes = Pedido.query.filter_by(estado='pendiente').count()
    pedidos_recientes_pendientes = Pedido.query.filter(
        Pedido.estado == 'pendiente',
        Pedido.fecha_pedido >= datetime.now(timezone.utc) - timedelta(days=1)
    ).count()
    
    hoy = datetime.now(timezone.utc).date()
    nuevas_reviews = Review.query.filter(
        func.date(Review.fecha_creacion) == hoy
    ).count()
    
    promedio_reviews = db.session.query(
        func.avg(Review.calificacion)
    ).scalar() or 0
    
    nuevos_usuarios = Usuario.query.filter(
        Usuario.fecha_registro >= datetime.now(timezone.utc) - timedelta(days=7)
    ).count()
    
    ventas_totales = db.session.query(
        func.sum(Pedido.total)
    ).filter(
        Pedido.estado == 'completado'
    ).scalar() or 0
    
    return render_template('admin/notificaciones.html',
                         productos=productos_bajo_stock,
                         pedidos_pendientes=pedidos_pendientes,
                         pedidos_recientes_pendientes=pedidos_recientes_pendientes,
                         nuevas_reviews=nuevas_reviews,
                         promedio_reviews=promedio_reviews,
                         nuevos_usuarios=nuevos_usuarios,
                         ventas_totales=ventas_totales,
                         hoy=hoy)

@app.route('/admin/proveedores')
@admin_required
def admin_proveedores():
    search_query = request.args.get('search', '').strip()
    query = Proveedor.query
    
    if search_query:
        query = query.filter(
            or_(
                Proveedor.nombre.ilike(f'{search_query}%'),
                Proveedor.identificacion.ilike(f'{search_query}%')
            )
        )
    
    proveedores = query.order_by(Proveedor.nombre.asc()).all()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        proveedores_data = []
        for proveedor in proveedores:
            proveedores_data.append({
                'id': proveedor.id,
                'nombre': proveedor.nombre,
                'identificacion': proveedor.identificacion,
                'tipo': proveedor.tipo,
                'email': proveedor.email,
                'telefono': proveedor.telefono,
                'activo': proveedor.activo
            })
        return jsonify({'proveedores': proveedores_data})
    
    return render_template('admin/proveedores.html', proveedores=proveedores)

@app.route('/admin/proveedores/agregar', methods=['GET', 'POST'])
@admin_required
def agregar_proveedor():
    if request.method == 'POST':
        nombre = request.form['nombre']
        identificacion = request.form['identificacion']
        tipo = request.form['tipo']
        direccion = request.form.get('direccion')
        telefono = request.form.get('telefono')
        email = request.form.get('email')

        if not nombre or not identificacion or not tipo:
            flash('Nombre, identificación y tipo son obligatorios', 'danger')
            return redirect(url_for('agregar_proveedor'))

        if Proveedor.query.filter_by(identificacion=identificacion).first():
            flash('La identificación ya está registrada', 'danger')
            return redirect(url_for('agregar_proveedor'))

        nuevo_proveedor = Proveedor(
            nombre=nombre,
            identificacion=identificacion,
            tipo=tipo,
            direccion=direccion,
            telefono=telefono,
            email=email,
            activo=True
        )
        
        try:
            db.session.add(nuevo_proveedor)
            db.session.commit()
            flash('Proveedor agregado con éxito', 'success')
            return redirect(url_for('admin_proveedores'))
        except IntegrityError:
            db.session.rollback()
            flash('Error al agregar el proveedor', 'danger')
            return redirect(url_for('agregar_proveedor'))

    return render_template('admin/agregar_proveedor.html')

@app.route('/admin/proveedores/editar/<int:proveedor_id>', methods=['GET', 'POST'])
@admin_required
def editar_proveedor(proveedor_id):
    proveedor = Proveedor.query.get_or_404(proveedor_id)
    
    if request.method == 'POST':
        proveedor.nombre = request.form['nombre']
        proveedor.identificacion = request.form['identificacion']
        proveedor.tipo = request.form['tipo']
        proveedor.direccion = request.form.get('direccion')
        proveedor.telefono = request.form.get('telefono')
        proveedor.email = request.form.get('email')

        if not proveedor.nombre or not proveedor.identificacion or not proveedor.tipo:
            flash('Nombre, identificación y tipo son obligatorios', 'danger')
            return redirect(url_for('editar_proveedor', proveedor_id=proveedor.id))

        try:
            db.session.commit()
            flash('Proveedor actualizado con éxito', 'success')
            return redirect(url_for('admin_proveedores'))
        except IntegrityError:
            db.session.rollback()
            flash('Error al actualizar el proveedor', 'danger')
            return redirect(url_for('editar_proveedor', proveedor_id=proveedor.id))

    return render_template('admin/editar_proveedor.html', proveedor=proveedor)

@app.route('/admin/proveedores/eliminar/<int:proveedor_id>', methods=['POST'])
@admin_required
def eliminar_proveedor(proveedor_id):
    proveedor = Proveedor.query.get_or_404(proveedor_id)
    
    try:
        # 1. Desasociar productos de la tienda (no eliminarlos)
        productos_tienda = Producto.query.filter_by(proveedor_id=proveedor_id).all()
        for producto in productos_tienda:
            producto.proveedor_id = None
            db.session.add(producto)
        
        # 2. Eliminar productos del proveedor (solo los que no están en compras)
        productos_proveedor = ProductoProveedor.query.filter_by(proveedor_id=proveedor_id).all()
        for producto in productos_proveedor:
            # Verificar si el producto está en alguna compra
            tiene_compras = DetalleCompra.query.filter_by(producto_id=producto.id).count() > 0
            if not tiene_compras:
                db.session.delete(producto)
        
        # 3. Eliminar compras asociadas (conservando los detalles para historial)
        compras = Compra.query.filter_by(proveedor_id=proveedor_id).all()
        for compra in compras:
            db.session.delete(compra)
        
        # 4. Finalmente eliminar el proveedor
        db.session.delete(proveedor)
        db.session.commit()
        
        flash('Proveedor eliminado correctamente. Los productos en tienda se han conservado pero desasociados.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar el proveedor: {str(e)}', 'danger')
    
    return redirect(url_for('admin_proveedores'))


@app.route('/admin/proveedores/<int:proveedor_id>/productos/editar/<int:producto_id>', methods=['GET', 'POST'])
@admin_required
def editar_producto_proveedor(proveedor_id, producto_id):
    proveedor = Proveedor.query.get_or_404(proveedor_id)
    producto = ProductoProveedor.query.filter_by(id=producto_id, proveedor_id=proveedor_id).first_or_404()
    producto_proveedor = ProductoProveedor.query.get_or_404(producto_id)
    
    # Actualizar producto en la tienda si existe
    producto_tienda = Producto.query.filter_by(
        nombre=producto_proveedor.nombre,
        proveedor_id=producto_proveedor.proveedor_id
    ).first()
    
    if producto_tienda:
        producto_tienda.nombre = request.form.get('nombre', producto_tienda.nombre)
        producto_tienda.descripcion = request.form.get('descripcion', producto_tienda.descripcion)
        producto_tienda.precio = float(request.form.get('precio', producto_tienda.precio))
        producto_tienda.stock = int(request.form.get('stock', producto_tienda.stock))
        producto_tienda.imagen = request.form.get('imagen', producto_tienda.imagen)
        
        db.session.commit()

    if request.method == 'POST':
        try:
            producto.nombre = request.form['nombre']
            producto.descripcion = request.form['descripcion']
            producto.marca = request.form.get('marca')
            producto.costo = float(request.form['costo'])
            producto.stock = int(request.form['stock'])
            
            db.session.commit()
            flash('Producto actualizado con éxito', 'success')
            return redirect(url_for('listar_productos_proveedor', proveedor_id=proveedor_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar el producto: {str(e)}', 'danger')

    return render_template('admin/editar_producto_proveedor.html', 
                         proveedor=proveedor, 
                         producto=producto)

@app.route('/admin/proveedores/<int:proveedor_id>/productos/eliminar/<int:producto_id>', methods=['POST'])
@admin_required
def eliminar_producto_proveedor(proveedor_id, producto_id):
    producto = ProductoProveedor.query.get_or_404(producto_id)
    try:
        # Eliminar también el producto de la tienda si existe
        producto_tienda = Producto.query.filter_by(
            nombre=producto.nombre,
            proveedor_id=producto.proveedor_id
        ).first()
        if producto_tienda:
            db.session.delete(producto_tienda)
        db.session.delete(producto)
        db.session.commit()
        flash('Producto eliminado correctamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar el producto: {str(e)}', 'danger')
    return redirect(url_for('listar_productos_proveedor', proveedor_id=proveedor_id))

@app.route('/admin/productos_proveedor/toggle_activo/<int:producto_id>', methods=['POST'])
@admin_required
def toggle_activo_producto_proveedor(producto_id):
    try:
        producto = ProductoProveedor.query.get_or_404(producto_id)
        data = request.get_json()
        nuevo_estado = data.get('activo', False)
        
        producto.activo = nuevo_estado
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Estado del producto actualizado a {"activo" if nuevo_estado else "inactivo"}'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al actualizar el estado: {str(e)}'
        }), 500

@app.route('/admin/proveedores/toggle_activo/<int:proveedor_id>', methods=['POST'])
@admin_required
def toggle_activo_proveedor(proveedor_id):
    try:
        proveedor = Proveedor.query.get_or_404(proveedor_id)
        data = request.get_json()
        nuevo_estado = data.get('activo', False)
        
        proveedor.activo = nuevo_estado
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Estado del proveedor actualizado a {"activo" if nuevo_estado else "inactivo"}'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al actualizar el estado: {str(e)}'
        }), 500

@app.route('/admin/proveedores/<int:proveedor_id>/productos')
@admin_required
def listar_productos_proveedor(proveedor_id):
    proveedor = Proveedor.query.get_or_404(proveedor_id)
    productos = ProductoProveedor.query.filter_by(proveedor_id=proveedor_id).all()
    
    return render_template('admin/productos_proveedor.html', 
                         proveedor=proveedor, 
                         productos=productos)

# En app.py, modificar la ruta /admin/ventas/agregar_cliente
@app.route('/admin/ventas/agregar_cliente', methods=['POST'])
@admin_required
def agregar_cliente_venta():
    try:
        data = request.get_json()
        
        # Validaciones básicas
        if not data.get('nombre') or not data.get('identificacion'):
            return jsonify({
                'success': False,
                'message': 'Nombre e identificación son obligatorios'
            }), 400

        # Verificar si el usuario ya existe
        if Usuario.query.filter_by(identificacion=data['identificacion']).first():
            return jsonify({
                'success': False,
                'message': 'El número de identificación ya está registrado'
            }), 400

        if data.get('email') and Usuario.query.filter_by(email=data['email']).first():
            return jsonify({
                'success': False,
                'message': 'El correo electrónico ya está registrado'
            }), 400

        # Crear nuevo usuario/cliente con 100 puntos iniciales
        nuevo_cliente = Usuario(
            nombre=data['nombre'],
            identificacion=data['identificacion'],
            email=data.get('email'),
            password=generate_password_hash(data.get('identificacion')),  # Password por defecto = identificación
            es_admin=False,
            fecha_registro=datetime.now(timezone.utc),
            activo=True,
            puntos=100  # Puntos iniciales para nuevos clientes
        )
        
        db.session.add(nuevo_cliente)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Cliente registrado con éxito. Se han asignado 100 puntos de bienvenida.',
            'cliente': {
                'id': nuevo_cliente.id,
                'nombre': nuevo_cliente.nombre,
                'identificacion': nuevo_cliente.identificacion,
                'email': nuevo_cliente.email,
                'puntos': nuevo_cliente.puntos
            }
        })
        
    except IntegrityError as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': 'Error de integridad en la base de datos'
        }), 500
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al registrar el cliente: {str(e)}'
        }), 500



@app.route('/admin/proveedores/<int:proveedor_id>/productos/agregar', methods=['GET', 'POST'])
@admin_required
def agregar_producto_proveedor(proveedor_id):
    proveedor = Proveedor.query.get_or_404(proveedor_id)
    if not proveedor.activo:
        flash('No se pueden agregar productos a un proveedor inactivo', 'danger')
        return redirect(url_for('admin_proveedores'))
    
    if request.method == 'POST':
        try:
            nombre = request.form['nombre'].strip()
            descripcion = request.form['descripcion'].strip()
            marca = request.form.get('marca', '').strip() or None
            costo = float(request.form['costo'])
            stock = int(request.form['stock'])
            
            if not nombre or not descripcion:
                flash('El nombre y la descripción son obligatorios', 'danger')
                return redirect(url_for('agregar_producto_proveedor', proveedor_id=proveedor_id))
            
            if costo <= 0:
                flash('El costo debe ser mayor que cero', 'danger')
                return redirect(url_for('agregar_producto_proveedor', proveedor_id=proveedor_id))
            
            if stock < 0:
                flash('El stock no puede ser negativo', 'danger')
                return redirect(url_for('agregar_producto_proveedor', proveedor_id=proveedor_id))
            
            nuevo_producto = ProductoProveedor(
                proveedor_id=proveedor_id,
                nombre=nombre,
                descripcion=descripcion,
                marca=marca,
                costo=costo,
                stock=stock,
                activo=True
            )
            db.session.add(nuevo_producto)
            db.session.commit()
            flash('Producto agregado al proveedor con éxito', 'success')
            return redirect(url_for('listar_productos_proveedor', proveedor_id=proveedor_id))
        except ValueError:
            db.session.rollback()
            flash('Datos inválidos para costo o stock', 'danger')
        except IntegrityError:
            db.session.rollback()
            flash('Error al agregar el producto', 'danger')
        return redirect(url_for('agregar_producto_proveedor', proveedor_id=proveedor_id))
    
    return render_template('admin/agregar_producto_proveedor.html', proveedor=proveedor)



@app.route('/admin/compras')
@admin_required
def admin_compras():
    # Obtener parámetros de búsqueda
    proveedor_id = request.args.get('proveedor_id', type=int)
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    
    query = Compra.query
    
    # Aplicar filtros
    if proveedor_id:
        query = query.filter_by(proveedor_id=proveedor_id)
    
    try:
        if fecha_desde:
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
            query = query.filter(Compra.fecha_compra >= fecha_desde_dt)
        if fecha_hasta:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            query = query.filter(Compra.fecha_compra <= fecha_hasta_dt.replace(hour=23, minute=59, second=59))
    except ValueError:
        flash('Formato de fecha inválido. Use YYYY-MM-DD', 'danger')
        return redirect(url_for('admin_compras'))
    
    compras = query.order_by(Compra.fecha_compra.desc()).all()
    proveedores = Proveedor.query.filter_by(activo=True).order_by(Proveedor.nombre).all()
    
    # Si es una petición AJAX, devolver JSON
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        compras_data = [{
            'id': compra.id,
            'fecha': compra.fecha_compra.strftime('%d/%m/%Y %H:%M:%S'),
            'proveedor': compra.proveedor.nombre,
            'total': float(compra.total),
            'productos': sum(d.cantidad for d in compra.detalles),
            'observaciones': compra.observaciones or 'Sin observaciones'
        } for compra in compras]
        return jsonify({
            'compras': compras_data,
            'proveedores': [{'id': p.id, 'nombre': p.nombre} for p in proveedores]
        })
    
    return render_template('admin/productos_comprados.html', 
                         compras=compras,
                         proveedores=proveedores)

@app.route('/admin/ventas/vender')
@admin_required
def vender_productos():
    # Obtener datos necesarios
    categorias_principales = Categoria.query.filter_by(parent_id=None, activa=True).all()
    usuarios = Usuario.query.filter_by(es_admin=False, activo=True).all()
    
    return render_template('admin/vender.html', 
                         categorias_principales=categorias_principales,
                         usuarios=usuarios)

@app.route('/admin/productos/search')
@admin_required
def buscar_productos_admin():
    query = request.args.get('q', '').strip()
    category_id = request.args.get('category_id', type=int)
    stock_filter = request.args.get('stock', 'all')
    
    # Construir consulta
    productos_query = Producto.query.filter_by(activo=True)
    
    if query:
        productos_query = productos_query.filter(Producto.nombre.ilike(f'{query}%'))
    
    if category_id:
        productos_query = productos_query.filter_by(categoria_id=category_id)
    
    if stock_filter == 'in_stock':
        productos_query = productos_query.filter(Producto.stock > 0)
    elif stock_filter == 'low_stock':
        productos_query = productos_query.filter(Producto.stock > 0, Producto.stock < Producto.stock_minimo)
    
    productos = productos_query.limit(20).all()
    
    # Convertir a formato JSON
    productos_json = [{
        'id': p.id,
        'nombre': p.nombre,
        'descripcion': p.descripcion,
        'precio': float(p.precio),
        'imagen': p.imagen,
        'stock': p.stock,
        'descuento': p.descuento,
        'marca': p.marca,
        'categoria_id': p.categoria_id
    } for p in productos]
    
    return jsonify(productos_json)

@app.route('/admin/ventas/crear', methods=['POST'])
@admin_required
def crear_venta():
    try:
        if not request.is_json:
            return jsonify({'success': False, 'message': 'El contenido debe ser tipo JSON'}), 400

        data = request.get_json()
        
        # Validaciones básicas
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400

        if 'items' not in data or not isinstance(data['items'], list) or len(data['items']) == 0:
            return jsonify({'success': False, 'message': 'No hay productos en la venta'}), 400

        # Validar campos numéricos incluyendo puntos usados
        required_fields = ['subtotal', 'discounts', 'tax', 'total']
        for field in required_fields:
            if field not in data or not isinstance(data[field], (int, float)):
                return jsonify({
                    'success': False,
                    'message': f'El campo {field} es requerido y debe ser numérico'
                }), 400

        # Obtener información del cliente si existe
        cliente = None
        puntos_usados = 0
        if data.get('clientId'):
            cliente = Usuario.query.get(data['clientId'])
            if not cliente:
                return jsonify({'success': False, 'message': 'Cliente no encontrado'}), 400
            
            # Validar puntos usados si aplica
            if 'puntosUsados' in data and data['puntosUsados'] > 0:
                puntos_usados = int(data['puntosUsados'])
                if puntos_usados > cliente.puntos:
                    return jsonify({
                        'success': False,
                        'message': 'El cliente no tiene suficientes puntos'
                    }), 400
                
                # Calcular descuento por puntos (1 punto = $0.10 de descuento)
                descuento_por_puntos = puntos_usados * 0.10
                if descuento_por_puntos > data['total']:
                    return jsonify({
                        'success': False,
                        'message': 'El descuento por puntos no puede exceder el total de la venta'
                    }), 400

        # Calcular puntos ganados (10 punto por cada $100)
        puntos_ganados = int(float(data['total']) // 10)

        # Crear nuevo pedido
        nuevo_pedido = Pedido(
            usuario_id=data.get('clientId'),
            fecha_pedido=datetime.now(timezone.utc),
            total=float(data['total']),
            nombre=cliente.nombre if cliente else 'Cliente general',
            direccion=cliente.direccion if (cliente and hasattr(cliente, 'direccion')) else 'Tienda física',
            ciudad=cliente.ciudad if (cliente and hasattr(cliente, 'ciudad')) else 'Ciudad',
            codigo_postal=cliente.codigo_postal if (cliente and hasattr(cliente, 'codigo_postal')) else '000000',
            telefono=cliente.telefono if (cliente and hasattr(cliente, 'telefono')) else '0000000000',
            metodo_pago=data.get('paymentMethod', 'efectivo'),
            estado='completado',
            puntos_usados=puntos_usados,
            puntos_ganados=puntos_ganados
        )
        db.session.add(nuevo_pedido)
        db.session.flush()
        
        # Procesar cada item del carrito
        productos_sin_stock = []
        for item in data['items']:
            producto = Producto.query.get(item['productId'])
            if not producto:
                continue
            
            if producto.stock < item['quantity']:
                productos_sin_stock.append({
                    'nombre': producto.nombre,
                    'stock_disponible': producto.stock,
                    'stock_solicitado': item['quantity']
                })
                continue
            
            detalle = DetallePedido(
                pedido_id=nuevo_pedido.id,
                producto_id=producto.id,
                cantidad=item['quantity'],
                precio=float(item['price']),
                descuento_aplicado=float(item.get('discount', 0))
            )
            db.session.add(detalle)
            producto.stock -= item['quantity']
        
        if productos_sin_stock:
            db.session.rollback()
            error_message = 'No hay suficiente stock para: '
            error_message += ', '.join([f"{p['nombre']} (Disp: {p['stock_disponible']}, Req: {p['stock_solicitado']})" 
                            for p in productos_sin_stock])
            return jsonify({'success': False, 'message': error_message}), 400

        # Asignar puntos al cliente si existe
        if cliente:
            if puntos_usados > 0:
                cliente.puntos -= puntos_usados
            cliente.puntos += puntos_ganados
            
                    
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Venta registrada con éxito',
            'pedido_id': nuevo_pedido.id,
            'puntos_ganados': puntos_ganados,
            'puntos_restantes': cliente.puntos if cliente else 0,
            'puntos_usados': puntos_usados,
            'total': float(data['total']),
            'fecha': nuevo_pedido.fecha_pedido.isoformat()
        })
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error al registrar venta: {str(e)}')
        return jsonify({
            'success': False,
            'message': 'Error interno al procesar la venta',
            'error': str(e)
        }), 500

@app.route('/admin/compras/<int:compra_id>')
@admin_required
def admin_detalle_compra(compra_id):
    compra = Compra.query.get_or_404(compra_id)
    detalles = DetalleCompra.query.filter_by(compra_id=compra_id).all()
    
    return render_template('admin/detalle_compra.html', 
                         compra=compra, 
                         detalles=detalles)


@app.route('/admin/usuarios/clientes')
@admin_required
def obtener_clientes():
    clientes = Usuario.query.filter_by(es_admin=False, activo=True).all()
    clientes_data = [{
        'id': c.id,
        'nombre': c.nombre,
        'identificacion': c.identificacion,
        'email': c.email,
        'puntos': c.puntos
    } for c in clientes]
    return jsonify(clientes_data)


@app.route('/admin/productos_proveedor/<int:proveedor_id>')
@admin_required
def productos_proveedor(proveedor_id):
    productos = ProductoProveedor.query.filter_by(proveedor_id=proveedor_id, activo=True).all()
    return jsonify([{
        'id': p.id,
        'nombre': p.nombre,
        'marca': p.marca,
        'costo': p.costo,
        'stock': p.stock
    } for p in productos])

# Modificar el endpoint para obtener productos de proveedores
@app.route('/admin/proveedores/<int:proveedor_id>/products/json')
@admin_required
def productos_proveedor_json(proveedor_id):
    try:
        app.logger.info(f"Solicitando productos para proveedor_id: {proveedor_id}")
        proveedor = Proveedor.query.filter_by(id=proveedor_id, activo=True).first()
        if not proveedor:
            app.logger.warning(f"Proveedor no encontrado o inactivo: {proveedor_id}")
            return jsonify({
                'success': False,
                'message': 'Proveedor no encontrado o inactivo'
            }), 404

        productos = ProductoProveedor.query.filter_by(
            proveedor_id=proveedor_id,
            activo=True,
            es_compra=False
        ).all()

        app.logger.info(f"Productos encontrados para proveedor {proveedor_id}: {len(productos)}")
        productos_list = [
            {
                'id': p.id,
                'nombre': p.nombre,
                'descripcion': p.descripcion or '',
                'marca': p.marca or 'Sin Marca',
                'costo': float(p.costo),
                'stock': int(p.stock),
                'formatted': f"{p.nombre} ({p.marca or 'Sin Marca'}) - ${float(p.costo):.2f} - Stock: {p.stock}"
            }
            for p in productos
        ]

        return jsonify({
            'success': True,
            'productos': productos_list,
            'proveedor': {
                'id': proveedor.id,
                'nombre': proveedor.nombre,
                'tipo': proveedor.tipo
            }
        })
    except Exception as e:
        app.logger.error(f'Error al obtener productos para proveedor {proveedor_id}: {str(e)}')
        return jsonify({
            'success': False,
            'message': f'Error al obtener productos: {str(e)}'
        }), 500

# Nuevo endpoint para procesar compras
@app.route('/admin/compras/procesar', methods=['POST'])
@admin_required
def procesar_compra():
    try:
        proveedor_id = int(request.form['proveedor_id'])
        categoria_id = int(request.form['categoria_id'])
        observaciones = request.form.get('observaciones', '')

        logging.info(f"Form data received: {request.form}")

        categoria = db.session.get(Categoria, categoria_id)
        if not categoria or not categoria.activa:
            flash('Categoría no válida o inactiva', 'danger')
            return redirect(url_for('agregar_compra'))

        proveedor = db.session.get(Proveedor, proveedor_id)
        if not proveedor or not proveedor.activo:
            flash('Proveedor no válido o inactivo', 'danger')
            return redirect(url_for('agregar_compra'))

        subtotal = 0
        detalles = []

        product_count = int(request.form.get('product_count', 0))
        for index in range(product_count):
            producto_proveedor_id = int(request.form[f'producto_proveedor_id_{index}'])
            cantidad = int(request.form[f'cantidad_{index}'])
            costo_unitario = float(request.form[f'costo_unitario_{index}'])

            logging.info(f"Processing product {index}: ID={producto_proveedor_id}, Cantidad={cantidad}, Costo={costo_unitario}")

            if cantidad <= 0:
                flash('La cantidad debe ser mayor que cero', 'danger')
                return redirect(url_for('agregar_compra'))

            producto_proveedor = db.session.get(ProductoProveedor, producto_proveedor_id)
            if not producto_proveedor or producto_proveedor.proveedor_id != proveedor_id or not producto_proveedor.activo:
                flash(f'Producto no válido para el proveedor seleccionado', 'danger')
                return redirect(url_for('agregar_compra'))

            if cantidad > producto_proveedor.stock:
                flash(f'No hay suficiente stock para {producto_proveedor.nombre}. Disponible: {producto_proveedor.stock}', 'danger')
                return redirect(url_for('agregar_compra'))

            if costo_unitario <= 0:
                flash(f'El costo unitario para {producto_proveedor.nombre} debe ser mayor que cero', 'danger')
                return redirect(url_for('agregar_compra'))

            subtotal_detalle = cantidad * costo_unitario
            subtotal += subtotal_detalle

            logging.info(f"Subtotal after product {index}: {subtotal}")

            detalle = DetalleCompra(
                producto_id=producto_proveedor_id,
                cantidad=cantidad,
                costo_unitario=costo_unitario
            )
            detalles.append(detalle)

        if not detalles:
            flash('Debe seleccionar al menos un producto para la compra', 'danger')
            return redirect(url_for('agregar_compra'))

        if subtotal <= 0:
            flash('El subtotal de la compra debe ser mayor que cero', 'danger')
            return redirect(url_for('agregar_compra'))

        iva = subtotal * IVA_RATE
        total = subtotal + iva

        logging.info(f"Final calculations: Subtotal={subtotal}, IVA={iva}, Total={total}")

        nueva_compra = Compra(
            proveedor_id=proveedor_id,
            subtotal=subtotal,
            total=total,
            fecha_compra=datetime.now(timezone.utc),
            observaciones=observaciones
        )
        db.session.add(nueva_compra)
        db.session.flush()

        for detalle in detalles:
            detalle.compra_id = nueva_compra.id
            db.session.add(detalle)

            producto_proveedor = db.session.get(ProductoProveedor, detalle.producto_id)
            producto_proveedor.stock -= detalle.cantidad

            producto_tienda = Producto.query.filter_by(
                nombre=producto_proveedor.nombre,
                marca=producto_proveedor.marca,
                categoria_id=categoria_id,
                proveedor_id=proveedor_id
            ).first()

            if producto_tienda:
                producto_tienda.stock = min(producto_tienda.stock + detalle.cantidad, producto_tienda.stock_maximo)
            else:
                nuevo_producto = Producto(
                    nombre=producto_proveedor.nombre,
                    descripcion=producto_proveedor.descripcion,
                    marca=producto_proveedor.marca,
                    precio=detalle.costo_unitario * 1.5,
                    stock=detalle.cantidad,
                    stock_minimo=10,
                    stock_maximo=100,
                    categoria_id=categoria_id,
                    proveedor_id=proveedor_id,
                    fecha_creacion=datetime.now(timezone.utc),
                    activo=True
                )
                db.session.add(nuevo_producto)

        db.session.commit()
        flash('Compra registrada con éxito', 'success')
        return redirect(url_for('admin_compras'))
    except ValueError as e:
        db.session.rollback()
        flash(f'Datos inválidos en el formulario: {str(e)}', 'danger')
    except IntegrityError as e:
        db.session.rollback()
        flash(f'Error al registrar la compra: {str(e)}', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'Error inesperado: {str(e)}', 'danger')
    return redirect(url_for('agregar_compra'))

@app.route('/admin/productos_proveedor/<int:producto_id>/json')
@admin_required
def producto_proveedor_json(producto_id):
    producto = ProductoProveedor.query.get_or_404(producto_id)
    return jsonify({
        'id': producto.id,
        'nombre': producto.nombre,
        'descripcion': producto.descripcion,
        'marca': producto.marca,
        'costo': float(producto.costo),
        'stock': int(producto.stock)
    })


@app.route('/admin/compras/disponibles')
@admin_required
def compras_disponibles():
    # Obtener compras con detalles no convertidos
    compras = db.session.query(
        Compra.id,
        Compra.fecha_compra,
        Proveedor.nombre.label('proveedor_nombre'),
        func.count(DetalleCompra.id).label('productos_pendientes'),
        func.sum(DetalleCompra.cantidad * DetalleCompra.costo_unitario).label('total_pendiente')
    ).join(
        Proveedor, Compra.proveedor_id == Proveedor.id
    ).join(
        DetalleCompra, DetalleCompra.compra_id == Compra.id
    ).filter(
        DetalleCompra.convertido == False,
        Proveedor.activo == True
    ).group_by(
        Compra.id, Compra.fecha_compra, Proveedor.nombre
    ).order_by(
        Compra.fecha_compra.desc()
    ).all()
    
    return render_template('admin/compras_disponibles.html', compras=compras)




@app.route('/admin/ventas/completadas', methods=['GET', 'POST'])
@admin_required
def ventas_completadas():
    try:
        app.logger.info("Iniciando endpoint ventas_completadas")
        # Obtener parámetros de filtrado
        fecha_inicio = request.args.get('fecha_inicio', '')
        fecha_fin = request.args.get('fecha_fin', '')
        cliente_id = request.args.get('cliente_id', type=int)
        metodo_pago = request.args.get('metodo_pago', '')
        orden = request.args.get('orden', 'fecha_desc')
        pagina = request.args.get('page', 1, type=int)
        
        app.logger.debug(f"Parámetros recibidos: fecha_inicio={fecha_inicio}, fecha_fin={fecha_fin}, cliente_id={cliente_id}, metodo_pago={metodo_pago}, orden={orden}, pagina={pagina}")

        # Validar fechas
        fecha_inicio_dt = None
        fecha_fin_dt = None
        
        if fecha_inicio and fecha_inicio.strip():
            try:
                fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
                app.logger.debug(f"Fecha inicio convertida: {fecha_inicio_dt}")
            except ValueError as e:
                app.logger.error(f"Error al convertir fecha_inicio: {str(e)}")
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify({'success': False, 'message': 'Formato de fecha inicio inválido. Use YYYY-MM-DD'}), 400
                flash('Formato de fecha inicio inválido. Use YYYY-MM-DD', 'danger')
                return redirect(url_for('ventas_completadas'))
        
        if fecha_fin and fecha_fin.strip():
            try:
                fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
                app.logger.debug(f"Fecha fin convertida: {fecha_fin_dt}")
            except ValueError as e:
                app.logger.error(f"Error al convertir fecha_fin: {str(e)}")
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify({'success': False, 'message': 'Formato de fecha fin inválido. Use YYYY-MM-DD'}), 400
                flash('Formato de fecha fin inválido. Use YYYY-MM-DD', 'danger')
                return redirect(url_for('ventas_completadas'))
        
        # Si no hay fechas, usar el último mes por defecto
        if not fecha_inicio and not fecha_fin:
            fecha_fin_dt = datetime.now(timezone.utc)
            fecha_inicio_dt = fecha_fin_dt - timedelta(days=30)
            fecha_inicio = fecha_inicio_dt.strftime('%Y-%m-%d')
            fecha_fin = fecha_fin_dt.strftime('%Y-%m-%d')
            app.logger.debug(f"Fechas por defecto: inicio={fecha_inicio}, fin={fecha_fin}")

        # Construir consulta para la tabla paginada
        query = Pedido.query.filter_by(estado='completado')
        
        # Aplicar filtros de fecha
        if fecha_inicio_dt:
            query = query.filter(Pedido.fecha_pedido >= fecha_inicio_dt)
        if fecha_fin_dt:
            query = query.filter(Pedido.fecha_pedido <= fecha_fin_dt.replace(hour=23, minute=59, second=59))
        
        # Aplicar otros filtros
        if cliente_id:
            query = query.filter_by(usuario_id=cliente_id)
        if metodo_pago:
            query = query.filter_by(metodo_pago=metodo_pago)
        
        # Aplicar orden
        if orden == 'fecha_asc':
            query = query.order_by(Pedido.fecha_pedido.asc())
        elif orden == 'total_desc':
            query = query.order_by(Pedido.total.desc())
        elif orden == 'total_asc':
            query = query.order_by(Pedido.total.asc())
        else:  # fecha_desc por defecto
            query = query.order_by(Pedido.fecha_pedido.desc())
        
        # Paginación
        por_pagina = 15
        paginacion = query.paginate(page=pagina, per_page=por_pagina, error_out=False)
        ventas = paginacion.items
        app.logger.debug(f"Paginación: página={pagina}, total={paginacion.total}, páginas={paginacion.pages}")

        # Si es una solicitud AJAX, devolver solo los datos de la tabla
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            ventas_data = []
            for venta in ventas:
                app.logger.debug(f"Procesando venta ID={venta.id}, fecha_pedido={venta.fecha_pedido}")
                ventas_data.append({
                    'id': venta.id,
                    'fecha_pedido': venta.fecha_pedido.strftime('%Y-%m-%d %H:%M:%S'),
                    'usuario': {
                        'nombre': venta.usuario.nombre if venta.usuario else None,
                        'identificacion': venta.usuario.identificacion if venta.usuario else None,
                        'puntos': venta.usuario.puntos if venta.usuario else None
                    },
                    'detalles': [{
                        'producto': {'nombre': detalle.producto.nombre},
                        'cantidad': detalle.cantidad
                    } for detalle in venta.detalles],
                    'total': float(venta.total),
                    'metodo_pago': venta.metodo_pago,
                    'puntos_usados': venta.puntos_usados
                })
            
            return jsonify({
                'success': True,
                'ventas': ventas_data,
                'paginacion': {
                    'total': paginacion.total,
                    'pages': paginacion.pages,
                    'page': paginacion.page
                }
            })

        # Para solicitudes no-AJAX, calcular estadísticas y datos de gráficos
        query_base = Pedido.query.filter_by(estado='completado')
        
        # Aplicar filtros de fecha a la consulta base
        if fecha_inicio_dt:
            query_base = query_base.filter(Pedido.fecha_pedido >= fecha_inicio_dt)
        if fecha_fin_dt:
            query_base = query_base.filter(Pedido.fecha_pedido <= fecha_fin_dt.replace(hour=23, minute=59, second=59))
        
        # Calcular estadísticas
        total_ventas = query_base.count()
        monto_total = query_base.with_entities(func.sum(Pedido.total)).scalar() or 0
        promedio_venta = monto_total / total_ventas if total_ventas > 0 else 0
        app.logger.debug(f"Estadísticas: total_ventas={total_ventas}, monto_total={monto_total}, promedio_venta={promedio_venta}")
        
        # Obtener métodos de pago disponibles
        metodos_pago_disponibles = db.session.query(Pedido.metodo_pago).filter_by(estado='completado').distinct().all()
        metodos_pago_disponibles = [m[0] for m in metodos_pago_disponibles]
        
        # Obtener clientes para el filtro
        clientes = Usuario.query.filter_by(es_admin=False, activo=True).all()
        
        # Datos para gráficos (ventas por día)
        ventas_por_dia = db.session.query(
            Pedido.fecha_pedido.label('fecha'),
            Pedido.total
        ).filter_by(estado='completado')

        if fecha_inicio_dt:
            ventas_por_dia = ventas_por_dia.filter(Pedido.fecha_pedido >= fecha_inicio_dt)
        if fecha_fin_dt:
            ventas_por_dia = ventas_por_dia.filter(Pedido.fecha_pedido <= fecha_fin_dt)

        # Obtener todos los registros
        ventas_por_dia = ventas_por_dia.all()

        # Agrupar por fecha en Python
        ventas_agrupadas = defaultdict(float)
        for venta in ventas_por_dia:
            fecha = venta.fecha.date()  # Truncar a fecha (sin hora)
            ventas_agrupadas[fecha] += venta.total

        # Ordenar por fecha
        ventas_por_dia = sorted(ventas_agrupadas.items(), key=lambda x: x[0])

        ventas_por_dia_labels = [fecha.strftime('%d/%m') for fecha, _ in ventas_por_dia]
        ventas_por_dia_data = [float(total) for _, total in ventas_por_dia]
        app.logger.debug(f"Datos gráficos: ventas_por_dia_labels={ventas_por_dia_labels[:5]}, ventas_por_dia_data={ventas_por_dia_data[:5]}")
        
        # Métodos de pago
        metodos_pago_data = db.session.query(
            Pedido.metodo_pago,
            func.sum(Pedido.total).label('total')
        ).filter_by(estado='completado')
        
        if fecha_inicio_dt:
            metodos_pago_data = metodos_pago_data.filter(Pedido.fecha_pedido >= fecha_inicio_dt)
        if fecha_fin_dt:
            metodos_pago_data = metodos_pago_data.filter(Pedido.fecha_pedido <= fecha_fin_dt)
            
        metodos_pago_data = metodos_pago_data.group_by(Pedido.metodo_pago).all()
        
        metodos_pago_labels = [m[0].capitalize() for m in metodos_pago_data]
        metodos_pago_data = [float(m[1]) for m in metodos_pago_data]
        app.logger.debug(f"Datos gráficos: metodos_pago_labels={metodos_pago_labels}, metodos_pago_data={metodos_pago_data}")
        
        return render_template('admin/ventas_completadas.html',
                              ventas=ventas,
                              paginacion=paginacion,
                              total_ventas=total_ventas,
                              monto_total=monto_total,
                              promedio_venta=promedio_venta,
                              metodos_pago=metodos_pago_disponibles,
                              clientes=clientes,
                              fecha_inicio=fecha_inicio,
                              fecha_fin=fecha_fin,
                              cliente_id=cliente_id,
                              metodo_pago=metodo_pago,
                              orden=orden,
                              ventas_por_dia_labels=ventas_por_dia_labels,
                              ventas_por_dia_data=ventas_por_dia_data,
                              metodos_pago_labels=metodos_pago_labels,
                              metodos_pago_data=metodos_pago_data)
        
    except Exception as e:
        app.logger.error(f'Error en ventas_completadas: {str(e)}')
        import traceback
        app.logger.error(f'Traceback: {traceback.format_exc()}')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'Error interno del servidor'}), 500
        flash('Error al cargar las ventas completadas', 'danger')
        return redirect(url_for('ventas_completadas'))





@app.route('/admin/productos_proveedor')
@admin_required
def admin_productos_proveedor():
    # Solo productos base (no compras)
    productos = ProductoProveedor.query.filter_by(es_compra=False).all()
    return render_template('admin/productos_proveedor.html', productos=productos)

@app.route('/admin/compras/disponibles/json')
@admin_required
def compras_disponibles_json():
    # Obtener detalles de compra no convertidos
    detalles = db.session.query(
        DetalleCompra.id,
        DetalleCompra.compra_id,
        DetalleCompra.cantidad,
        DetalleCompra.costo_unitario,
        ProductoProveedor.id.label('producto_id'),
        ProductoProveedor.nombre.label('producto_nombre'),
        ProductoProveedor.descripcion.label('producto_descripcion'),
        ProductoProveedor.marca.label('producto_marca')
    ).join(
        ProductoProveedor, ProductoProveedor.id == DetalleCompra.producto_id
    ).filter(
        DetalleCompra.convertido == False,
        ProductoProveedor.activo == True
    ).all()
    
    detalles_list = [{
        'id': d.id,
        'compra_id': d.compra_id,
        'producto_id': d.producto_id,
        'producto_nombre': d.producto_nombre,
        'producto_descripcion': d.producto_descripcion,
        'producto_marca': d.producto_marca,
        'costo_unitario': float(d.costo_unitario),
        'cantidad': int(d.cantidad)
    } for d in detalles]
    
    return jsonify(detalles_list)


@app.route('/admin/compras/agregar', methods=['GET', 'POST'])
@admin_required
def agregar_compra():
    categorias_principales = Categoria.query.filter_by(parent_id=None, activa=True).all()
    subcategorias = Categoria.query.filter(Categoria.parent_id != None, Categoria.activa == True).all()
    proveedores = Proveedor.query.filter_by(activo=True).all()

    if request.method == 'POST':
        proveedor_id = request.form.get('proveedor_id', type=int)
        categoria_id = request.form.get('categoria_id', type=int)
        observaciones = request.form.get('observaciones', '').strip()
        IVA_RATE = 0.19
        subtotal = 0.0
        detalles = []

        if not proveedor_id:
            flash('Debe seleccionar un proveedor válido', 'danger')
            return redirect(url_for('agregar_compra'))

        if not categoria_id:
            flash('Debe seleccionar una categoría válida', 'danger')
            return redirect(url_for('agregar_compra'))

        proveedor = Proveedor.query.get_or_404(proveedor_id)
        categoria = Categoria.query.get_or_404(categoria_id)

        try:
            producto_indices = request.form.getlist('producto_id')
            if not producto_indices:
                flash('Debe seleccionar al menos un producto', 'danger')
                return redirect(url_for('agregar_compra'))

            for index in producto_indices:
                producto_proveedor_id = request.form.get(f'producto_id_{index}', type=int)
                cantidad = request.form.get(f'cantidad_{index}', type=int)
                costo_unitario = request.form.get(f'costo_unitario_{index}', type=float)

                if not all([producto_proveedor_id, cantidad, costo_unitario]):
                    flash('Todos los campos de producto son obligatorios', 'danger')
                    return redirect(url_for('agregar_compra'))

                if cantidad <= 0:
                    flash('La cantidad debe ser mayor que cero', 'danger')
                    return redirect(url_for('agregar_compra'))

                if costo_unitario <= 0:
                    flash('El costo unitario debe ser mayor que cero', 'danger')
                    return redirect(url_for('agregar_compra'))

                producto_proveedor = ProductoProveedor.query.filter_by(
                    id=producto_proveedor_id,
                    proveedor_id=proveedor_id,
                    activo=True
                ).first()

                if not producto_proveedor:
                    flash('Producto seleccionado no válido', 'danger')
                    return redirect(url_for('agregar_compra'))

                if cantidad > producto_proveedor.stock:
                    flash(f'No hay suficiente stock para {producto_proveedor.nombre}. Disponible: {producto_proveedor.stock}', 'danger')
                    return redirect(url_for('agregar_compra'))

                subtotal_detalle = cantidad * costo_unitario
                subtotal += subtotal_detalle

                detalle = DetalleCompra(
                    producto_id=producto_proveedor_id,
                    cantidad=cantidad,
                    costo_unitario=costo_unitario
                )
                detalles.append(detalle)

            iva = subtotal * IVA_RATE
            total = subtotal + iva

            nueva_compra = Compra(
                proveedor_id=proveedor_id,
                subtotal=subtotal,
                total=total,
                fecha_compra=datetime.now(timezone.utc),
                observaciones=observaciones
            )
            db.session.add(nueva_compra)
            db.session.flush()

            for detalle in detalles:
                detalle.compra_id = nueva_compra.id
                db.session.add(detalle)

                producto_proveedor = ProductoProveedor.query.get(detalle.producto_id)
                producto_proveedor.stock -= detalle.cantidad

                producto_tienda = Producto.query.filter_by(
                    nombre=producto_proveedor.nombre,
                    marca=producto_proveedor.marca,
                    categoria_id=categoria_id,
                    proveedor_id=proveedor_id
                ).first()

                if producto_tienda:
                    producto_tienda.stock = min(producto_tienda.stock + detalle.cantidad, producto_tienda.stock_maximo)
                    producto_tienda.costo = detalle.costo_unitario
                else:
                    nuevo_producto = Producto(
                        nombre=producto_proveedor.nombre,
                        descripcion=producto_proveedor.descripcion or 'Sin descripción',
                        marca=producto_proveedor.marca,
                        precio=detalle.costo_unitario * 1.5,
                        costo=detalle.costo_unitario,
                        stock=detalle.cantidad,
                        stock_minimo=10,
                        stock_maximo=100,
                        categoria_id=categoria_id,
                        proveedor_id=proveedor_id,
                        fecha_creacion=datetime.now(timezone.utc),
                        activo=True
                    )
                    db.session.add(nuevo_producto)

            db.session.commit()
            flash('Compra registrada con éxito', 'success')
            return redirect(url_for('admin_compras'))

        except ValueError as e:
            db.session.rollback()
            flash(f'Error en los datos del formulario: {str(e)}', 'danger')
        except IntegrityError as e:
            db.session.rollback()
            flash(f'Error al registrar la compra: {str(e)}', 'danger')
        except Exception as e:
            db.session.rollback()
            flash(f'Error inesperado al procesar la compra: {str(e)}', 'danger')
        return redirect(url_for('agregar_compra'))

    return render_template('admin/agregar_compra.html',
                          categorias=categorias_principales,
                          subcategorias=subcategorias,
                          proveedores=proveedores)


@app.route('/admin/inventario/generar_excel')
@admin_required
def generar_reporte_inventario_excel():
    try:
        # Obtener datos
        productos = Producto.query.filter_by(activo=True).all()
        total_productos = len(productos)
        valor_total_base = sum(p.precio * p.stock for p in productos)
        valor_total_final = sum(p.precio_final * p.stock for p in productos)
        descuento_total = valor_total_base - valor_total_final
        
        # Crear libro de trabajo y hojas
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_inventario = wb.create_sheet("Inventario")

        # Hoja Resumen
        ws_resumen.merge_cells('A1:J1')
        ws_resumen['A1'] = "REPORTE DE INVENTARIO"
        ws_resumen['A1'].font = Font(bold=True, size=14)
        ws_resumen['A1'].alignment = Alignment(horizontal='center')

        ws_resumen.merge_cells('A2:J2')
        ws_resumen['A2'] = f"Generado el: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M:%S')}"
        ws_resumen['A2'].alignment = Alignment(horizontal='center')

        ws_resumen['A4'] = "RESUMEN GENERAL"
        ws_resumen['A4'].font = Font(bold=True)
        ws_resumen.append(["Concepto", "Valor"])
        ws_resumen.append(["Total Productos", total_productos])
        ws_resumen.append(["Valor Total Base", valor_total_base])
        ws_resumen.append(["Descuento Total", descuento_total])
        ws_resumen.append(["Valor Total Final", valor_total_final])

        # Formato numérico para la columna de valores
        for row in ws_resumen['B5:B8']:
            row[0].number_format = '$#,##0.00'

        # Ajustar ancho de columnas
        ws_resumen.column_dimensions['A'].width = 20
        ws_resumen.column_dimensions['B'].width = 15

        # Hoja Inventario
        ws_inventario.merge_cells('A1:J1')
        ws_inventario['A1'] = "DETALLE DE PRODUCTOS"
        ws_inventario['A1'].font = Font(bold=True)
        ws_inventario['A1'].alignment = Alignment(horizontal='center')

        headers = ["ID", "Producto", "Marca", "Categoría", "Subcategoría", 
                  "Stock", "Precio Base", "Precio Final", "Descuento", "Valor Total"]
        ws_inventario.append(headers)

        for producto in productos:
            categoria_principal = producto.categoria.parent.nombre if producto.categoria.parent else producto.categoria.nombre
            subcategoria = producto.categoria.nombre if producto.categoria.parent else "-"
            ws_inventario.append([
                producto.id,
                producto.nombre,
                producto.marca or "-",
                categoria_principal,
                subcategoria,
                producto.stock,
                producto.precio,
                producto.precio_final,
                producto.descuento if producto.descuento > 0 else 0,
                producto.precio_final * producto.stock
            ])

        # Formato numérico para columnas
        for col in ['F', 'G', 'H', 'J']:  # Stock, Precio Base, Precio Final, Valor Total
            for cell in ws_inventario[col + '2:' + col + str(len(productos) + 2)]:
                cell[0].number_format = '$#,##0.00'
        for cell in ws_inventario['I2:I' + str(len(productos) + 2)]:  # Descuento
            cell[0].number_format = '0%'

        # Ajustar ancho de columnas
        for col in ws_inventario.column_dimensions:
            ws_inventario.column_dimensions[col].width = 15

        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Guardar localmente para depuración
        with open("temp_inventario_2025-06-15_2247.xlsx", "wb") as f:
            output.seek(0)
            f.write(output.getvalue())

        # Verificar tamaño
        output.seek(0)
        app.logger.info(f"Tamaño del archivo Excel: {output.tell()} bytes")
        if output.tell() == 0:
            raise Exception("El archivo Excel generado está vacío")

        # Crear respuesta
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        filename = f"reporte_inventario_{datetime.now(timezone.utc).strftime('%Y-%m-%d_%H%M%S')}.xlsx"
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Length'] = output.tell()

        return response

    except Exception as e:
        app.logger.error(f'Error al generar Excel: {str(e)}')
        flash(f'Error al generar el reporte Excel: {str(e)}', 'danger')
        return redirect(url_for('reportes_inventario'))


@app.route('/admin/inventario/reportes')
@admin_required
def reportes_inventario():
    # Obtener todos los productos activos
    productos = Producto.query.filter_by(activo=True).all()
    
    # Calcular estadísticas
    total_productos = len(productos)
    valor_total_base = sum(p.precio * p.stock for p in productos)
    valor_total_final = sum(p.precio_final * p.stock for p in productos)
    descuento_total = valor_total_base - valor_total_final
    
    productos_bajo_stock = [p for p in productos if p.stock < p.stock_minimo]
    productos_sobre_stock = [p for p in productos if p.stock > p.stock_maximo]
    
    # Prepara datos para gráficos
    categorias_stock = {}
    categorias_valor = {}
    for p in productos:
        cat_nombre = p.categoria.nombre
        if cat_nombre not in categorias_stock:
            categorias_stock[cat_nombre] = {"stock": 0, "valor_base": 0, "valor_final": 0}
        
        categorias_stock[cat_nombre]["stock"] += p.stock
        categorias_stock[cat_nombre]["valor_base"] += p.precio * p.stock
        categorias_stock[cat_nombre]["valor_final"] += p.precio_final * p.stock
    
    # Obtener categorías principales con sus subcategorías
    categorias_principales = Categoria.query.filter_by(parent_id=None, activa=True).all()
    categorias_con_sub = []
    
    for cat in categorias_principales:
        subcategorias = Categoria.query.filter_by(parent_id=cat.id, activa=True).all()
        categorias_con_sub.append({
            'principal': cat,
            'subcategorias': subcategorias
        })
    
    # Calcular el porcentaje de descuento promedio
    descuento_promedio = 0
    productos_con_descuento = [p for p in productos if p.descuento > 0]
    if productos_con_descuento:
        descuento_promedio = sum(p.descuento for p in productos_con_descuento) / len(productos_con_descuento)
    
    return render_template('admin/reportes_inventario.html', 
                           productos=productos,
                           total_productos=total_productos,
                           valor_total_base=valor_total_base,
                           valor_total_final=valor_total_final,
                           descuento_total=descuento_total,
                           productos_bajo_stock=len(productos_bajo_stock),
                           productos_sobre_stock=len(productos_sobre_stock),
                           categorias_stock=categorias_stock,
                           categorias_con_sub=categorias_con_sub,
                           descuento_promedio=descuento_promedio)




def generar_pdf(html):
    result = BytesIO()
    pisa.CreatePDF(BytesIO(html.encode('utf-8')), result)
    return result.getvalue()

def image_to_base64(image_path):
    with open(image_path, "rb") as image_file:
        return "data:image/png;base64," + base64.b64encode(image_file.read()).decode('utf-8')

# Añadir esta ruta para la página de descargas
@app.route('/admin/descargas')
@admin_required
def admin_descargas():
    return render_template('admin/descargas.html')


@app.route('/admin/inventario/generar_pdf')
@admin_required
def generar_reporte_inventario_pdf():
    try:
        # Obtener todos los productos activos
        productos = Producto.query.filter_by(activo=True).all()
        
        # Calcular estadísticas
        total_productos = len(productos)
        valor_total_base = sum(p.precio * p.stock for p in productos)
        valor_total_final = sum(p.precio_final * p.stock for p in productos)
        descuento_total = valor_total_base - valor_total_final
        
        productos_bajo_stock = len([p for p in productos if p.stock < p.stock_minimo])
        productos_sobre_stock = len([p for p in productos if p.stock > p.stock_maximo])
        
        # Preparar datos para gráficos
        categorias_stock = {}
        for p in productos:
            cat_nombre = p.categoria.nombre
            if cat_nombre not in categorias_stock:
                categorias_stock[cat_nombre] = {"stock": 0, "valor_base": 0, "valor_final": 0}
            
            categorias_stock[cat_nombre]["stock"] += p.stock
            categorias_stock[cat_nombre]["valor_base"] += p.precio * p.stock
            categorias_stock[cat_nombre]["valor_final"] += p.precio_final * p.stock
        
        # Calcular el porcentaje de descuento promedio
        descuento_promedio = 0
        productos_con_descuento = [p for p in productos if p.descuento > 0]
        if productos_con_descuento:
            descuento_promedio = sum(p.descuento for p in productos_con_descuento) / len(productos_con_descuento)
        
        # Renderizar el template HTML
        fecha = datetime.now(timezone.utc).strftime("%d/%m/%Y")
        html = render_template('admin/reporte_inventario_pdf.html', 
                             fecha=fecha,
                             productos=productos,
                             total_productos=total_productos,
                             valor_total_base=valor_total_base,
                             valor_total_final=valor_total_final,
                             descuento_total=descuento_total,
                             productos_bajo_stock=productos_bajo_stock,
                             productos_sobre_stock=productos_sobre_stock,
                             categorias_stock=categorias_stock,
                             descuento_promedio=descuento_promedio)
        
        pdf = generar_pdf(html)
        # Crear respuesta
        response = make_response(pdf)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=reporte_inventario_{datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")}.pdf'
        
        return response
        
    except Exception as e:
        app.logger.error(f'Error al generar PDF: {str(e)}')
        flash(f'Error al generar el reporte PDF: {str(e)}', 'danger')
        return redirect(url_for('reportes_inventario'))


# Reporte de Productos
@app.route('/admin/descargas/productos')
@admin_required
def descargar_productos_pdf():
    productos = Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()
    
    # Calcular estadísticas
    total_productos = len(productos)
    productos_con_descuento = len([p for p in productos if p.descuento > 0])
    valor_inventario = sum(p.precio_final * p.stock for p in productos)
    
    html = render_template('admin/reportes/productos_pdf.html', 
                         productos=productos,
                         total_productos=total_productos,
                         productos_con_descuento=productos_con_descuento,
                         valor_inventario=valor_inventario,
                         fecha=datetime.now(timezone.utc).strftime("%d/%m/%Y"),
                         logo=image_to_base64(os.path.join(app.root_path, 'static', 'img', 'logo2.png')))
    
    pdf = generar_pdf(html)
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=reporte_productos.pdf'
    return response

# Reporte de Usuarios
@app.route('/admin/descargas/usuarios')
@admin_required
def descargar_usuarios_pdf():
    usuarios = Usuario.query.order_by(Usuario.nombre).all()
    
    # Estadísticas
    total_usuarios = len(usuarios)
    admins = len([u for u in usuarios if u.es_admin])
    clientes_activos = len([u for u in usuarios if not u.es_admin and u.activo])
    
    html = render_template('admin/reportes/usuarios_pdf.html', 
                         usuarios=usuarios,
                         total_usuarios=total_usuarios,
                         admins=admins,
                         clientes_activos=clientes_activos,
                         fecha=datetime.now(timezone.utc).strftime("%d/%m/%Y"),
                         logo=image_to_base64(os.path.join(app.root_path, 'static', 'img', 'logo2.png')))
    
    pdf = generar_pdf(html)
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=reporte_usuarios.pdf'
    return response

# Reporte de Compras
@app.route('/admin/descargas/compras')
@admin_required
def descargar_compras_pdf():
    compras = Compra.query.order_by(Compra.fecha_compra.desc()).all()
    
    # Estadísticas
    total_compras = len(compras)
    total_gastado = sum(c.total for c in compras)
    proveedores = {c.proveedor.nombre for c in compras}
    
    html = render_template('admin/reportes/compras_pdf.html', 
                         compras=compras,
                         total_compras=total_compras,
                         total_gastado=total_gastado,
                         proveedores=len(proveedores),
                         fecha=datetime.now(timezone.utc).strftime("%d/%m/%Y"),
                         logo=image_to_base64(os.path.join(app.root_path, 'static', 'img', 'logo2.png')))
    
    pdf = generar_pdf(html)
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=reporte_compras.pdf'
    return response

# Reporte de Ventas
@app.route('/admin/descargas/ventas')
@admin_required
def descargar_ventas_pdf():
    pedidos = Pedido.query.order_by(Pedido.fecha_pedido.desc()).all()
    
    # Estadísticas
    total_pedidos = len(pedidos)
    total_ventas = sum(p.total for p in pedidos)
    pedidos_completados = len([p for p in pedidos if p.estado == 'completado'])
    
    html = render_template('admin/reportes/ventas_pdf.html', 
                         pedidos=pedidos,
                         total_pedidos=total_pedidos,
                         total_ventas=total_ventas,
                         pedidos_completados=pedidos_completados,
                         fecha=datetime.now(timezone.utc).strftime("%d/%m/%Y"),
                         logo=image_to_base64(os.path.join(app.root_path, 'static', 'img', 'logo2.png')))
    
    pdf = generar_pdf(html)
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=reporte_ventas.pdf'
    return response

# Añadir estos nuevos endpoints después de los existentes

# Reporte de Proveedores
@app.route('/admin/descargas/proveedores')
@admin_required
def descargar_proveedores_pdf():
    proveedores = Proveedor.query.order_by(Proveedor.nombre).all()
    
    # Estadísticas - Calcular todo en Python antes de pasar a la plantilla
    total_proveedores = len(proveedores)
    proveedores_activos = len([p for p in proveedores if p.activo])
    productos_por_proveedor = {p.nombre: p.productos.count() for p in proveedores}
    total_productos = sum(productos_por_proveedor.values())  # Calculamos aquí el total
    
    html = render_template('admin/reportes/proveedores_pdf.html', 
                         proveedores=proveedores,
                         total_proveedores=total_proveedores,
                         proveedores_activos=proveedores_activos,
                         productos_por_proveedor=productos_por_proveedor,
                         total_productos=total_productos,  # Pasamos el total ya calculado
                         fecha=datetime.now(timezone.utc).strftime("%d/%m/%Y"),
                         logo=image_to_base64(os.path.join(app.root_path, 'static', 'img', 'logo2.png')))
    
    pdf = generar_pdf(html)
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=reporte_proveedores.pdf'
    return response
    return response

# Reporte de Categorías
@app.route('/admin/descargas/categorias')
@admin_required
def descargar_categorias_pdf():
    # Obtener categorías principales
    categorias_principales = Categoria.query.filter_by(parent_id=None, activa=True).all()
    
    categorias_con_datos = []
    total_subcategorias = 0
    total_productos = 0
    
    for cat in categorias_principales:
        # Obtener subcategorías - convertir a lista
        subcategorias = cat.subcategorias.all()  # Correcto: subcategorias es una consulta dinámica
        total_subcategorias += len(subcategorias)
        
        # Procesar subcategorías y sus productos
        subcategorias_con_datos = []
        productos_subcategorias = 0
        
        for subcat in subcategorias:
            # Obtener productos - ya es una lista
            productos = subcat.productos  # Eliminado .all(), ya es una lista
            productos_subcategorias += len(productos)
            subcategorias_con_datos.append({
                'subcategoria': subcat,
                'productos': productos,
                'total_productos': len(productos)
            })
        
        # Productos directos en la categoría principal
        productos_principal = cat.productos  # Eliminado .all(), ya es una lista
        total_productos_principal = len(productos_principal)
        
        # Totales para esta categoría
        total_categoria = total_productos_principal + productos_subcategorias
        total_productos += total_categoria
        
        categorias_con_datos.append({
            'categoria': cat,
            'subcategorias': subcategorias_con_datos,
            'productos_principal': productos_principal,
            'total_productos_principal': total_productos_principal,
            'total_productos_subcategorias': productos_subcategorias,
            'total_productos': total_categoria
        })
    
    html = render_template('admin/reportes/categorias_pdf.html', 
                         categorias=categorias_con_datos,
                         total_categorias=len(categorias_principales),
                         total_subcategorias=total_subcategorias,
                         total_productos=total_productos,
                         fecha=datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M"),
                         logo=image_to_base64(os.path.join(app.root_path, 'static', 'img', 'logo2.png')))
    
    pdf = generar_pdf(html)
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=reporte_categorias.pdf'
    return response


@app.route('/admin/descargas/productos_excel')
@admin_required
def descargar_productos_excel():
    try:
        productos = Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()
        
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet('Productos')
        
        # Formatos
        header_format = workbook.add_format({
            'bold': True, 'bg_color': '#4472C4', 'font_color': 'white',
            'border': 1, 'align': 'center'
        })
        currency_format = workbook.add_format({'num_format': '$#,##0.00'})
        percent_format = workbook.add_format({'num_format': '0.00%'})
        
        # Encabezados
        headers = [
            'ID', 'Nombre', 'Marca', 'Categoría', 'Subcategoría',
            'Precio', 'Descuento', 'Precio Final', 'Stock',
            'Stock Mínimo', 'Stock Máximo', 'Destacado', 'Proveedor'
        ]
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)
        
        # Datos
        for row_num, producto in enumerate(productos, start=1):
            categoria = producto.categoria
            categoria_principal = categoria.parent.nombre if categoria and categoria.parent else (categoria.nombre if categoria else 'N/A')
            subcategoria = categoria.nombre if categoria and categoria.parent else 'N/A'
            marca = producto.marca or 'N/A'
            proveedor = producto.proveedor.nombre if producto.proveedor else 'N/A'
            precio_final = producto.precio * (1 - producto.descuento/100) if producto.descuento > 0 else producto.precio
            
            data = [
                producto.id,
                producto.nombre,
                marca,
                categoria_principal,
                subcategoria,
                producto.precio,
                producto.descuento/100 if producto.descuento > 0 else 0,
                precio_final,
                producto.stock,
                producto.stock_minimo,
                producto.stock_maximo,
                'Sí' if producto.destacado else 'No',
                proveedor
            ]
            
            for col_num, value in enumerate(data):
                if col_num in [5, 7]:  # Precio, Precio Final
                    worksheet.write_number(row_num, col_num, value, currency_format)
                elif col_num == 6:  # Descuento
                    worksheet.write_number(row_num, col_num, value, percent_format)
                elif col_num in [8, 9, 10]:  # Stocks
                    worksheet.write_number(row_num, col_num, value)
                else:
                    worksheet.write(row_num, col_num, value)
        
        # Ajustar anchos
        worksheet.set_column('A:A', 8)   # ID
        worksheet.set_column('B:B', 30)  # Nombre
        worksheet.set_column('C:C', 20)  # Marca
        worksheet.set_column('D:D', 20)  # Categoría
        worksheet.set_column('E:E', 20)  # Subcategoría
        worksheet.set_column('F:H', 15)  # Precios
        worksheet.set_column('I:K', 12)  # Stocks
        worksheet.set_column('L:L', 10)  # Destacado
        worksheet.set_column('M:M', 25)  # Proveedor
        
        workbook.close()
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=reporte_productos_{datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        app.logger.error(f'Error al generar Excel de productos: {str(e)}', exc_info=True)
        flash(f'Error al generar el reporte de productos en Excel: {str(e)}', 'danger')
        return redirect(url_for('admin_productos'))


@app.route('/admin/descargas/categorias_excel')
@admin_required
def descargar_categorias_excel():
    try:
        categorias_principales = Categoria.query.filter_by(parent_id=None, activa=True).all()
        
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet('Categorías')
        
        # Formatos
        header_format = workbook.add_format({
            'bold': True, 'bg_color': '#4472C4', 'font_color': 'white',
            'border': 1, 'align': 'center'
        })
        main_cat_format = workbook.add_format({
            'bold': True, 'bg_color': '#E6E6E6'
        })
        
        # Encabezados
        headers = [
            'ID', 'Nombre', 'Tipo', 'Activa', 'Productos',
            'Subcategorías'
        ]
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)
        
        # Datos
        row_num = 1
        for categoria in categorias_principales:
            subcategorias = Categoria.query.filter_by(parent_id=categoria.id, activa=True).all()
            productos_count = Producto.query.filter_by(categoria_id=categoria.id, activo=True).count()
            
            # Categoría principal
            worksheet.write(row_num, 0, categoria.id, main_cat_format)
            worksheet.write(row_num, 1, categoria.nombre, main_cat_format)
            worksheet.write(row_num, 2, 'Principal', main_cat_format)
            worksheet.write(row_num, 3, 'Sí' if categoria.activa else 'No', main_cat_format)
            worksheet.write(row_num, 4, productos_count, main_cat_format)
            worksheet.write(row_num, 5, len(subcategorias), main_cat_format)
            row_num += 1
            
            # Subcategorías
            for subcat in subcategorias:
                subcat_productos_count = Producto.query.filter_by(categoria_id=subcat.id, activo=True).count()
                worksheet.write(row_num, 0, subcat.id)
                worksheet.write(row_num, 1, subcat.nombre)
                worksheet.write(row_num, 2, 'Subcategoría')
                worksheet.write(row_num, 3, 'Sí' if subcat.activa else 'No')
                worksheet.write(row_num, 4, subcat_productos_count)
                worksheet.write(row_num, 5, 'N/A')
                row_num += 1
        
        # Ajustar anchos
        worksheet.set_column('A:A', 8)   # ID
        worksheet.set_column('B:B', 30)  # Nombre
        worksheet.set_column('C:C', 15)  # Tipo
        worksheet.set_column('D:D', 10)  # Activa
        worksheet.set_column('E:E', 10)  # Productos
        worksheet.set_column('F:F', 12)  # Subcategorías
        
        workbook.close()
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=reporte_categorias_{datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        app.logger.error(f'Error al generar Excel de categorías: {str(e)}', exc_info=True)
        flash(f'Error al generar el reporte de categorías en Excel: {str(e)}', 'danger')
        return redirect(url_for('admin_categorias'))



@app.route('/admin/descargas/compras_excel')
@admin_required
def descargar_compras_excel():
    try:
        compras = Compra.query.order_by(Compra.fecha_compra.desc()).all()
        
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet('Compras')
        
        # Formatos
        header_format = workbook.add_format({
            'bold': True, 'bg_color': '#4472C4', 'font_color': 'white', 
            'border': 1, 'align': 'center'
        })
        date_format = workbook.add_format({'num_format': 'dd/mm/yyyy'})
        currency_format = workbook.add_format({'num_format': '$#,##0.00'})
        
        # Encabezados
        headers = [
            'ID', 'Fecha', 'Proveedor', 'Total', 'Productos', 
            'Observaciones'
        ]
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)
        
        # Datos
        for row_num, compra in enumerate(compras, start=1):
            total_productos = sum(d.cantidad for d in compra.detalles)
            
            data = [
                compra.id,
                compra.fecha_compra,
                compra.proveedor.nombre if compra.proveedor else 'N/A',
                compra.total,
                total_productos,
                compra.observaciones or 'N/A'
            ]
            
            for col_num, value in enumerate(data):
                if col_num == 1:  # Fecha
                    worksheet.write_datetime(row_num, col_num, compra.fecha_compra, date_format)
                elif col_num == 3:  # Total
                    worksheet.write_number(row_num, col_num, value, currency_format)
                else:
                    worksheet.write(row_num, col_num, value)
        
        # Ajustar anchos
        worksheet.set_column('A:A', 8)   # ID
        worksheet.set_column('B:B', 15)  # Fecha
        worksheet.set_column('C:C', 30)  # Proveedor
        worksheet.set_column('D:D', 15)  # Total
        worksheet.set_column('E:E', 12)  # Productos
        worksheet.set_column('F:F', 40)  # Observaciones
        
        # Hoja de Detalles
        details_sheet = workbook.add_worksheet('Detalles')
        
        # Encabezados detalles
        detail_headers = [
            'Compra ID', 'Producto', 'Proveedor', 'Cantidad', 
            'Costo Unitario', 'Subtotal'
        ]
        for col_num, header in enumerate(detail_headers):
            details_sheet.write(0, col_num, header, header_format)
        
        # Datos detalles
        row_num = 1
        for compra in compras:
            for detalle in compra.detalles:
                producto = ProductoProveedor.query.get(detalle.producto_id)
                
                details_sheet.write(row_num, 0, compra.id)
                details_sheet.write(row_num, 1, producto.nombre)
                details_sheet.write(row_num, 2, compra.proveedor.nombre if compra.proveedor else 'N/A')
                details_sheet.write_number(row_num, 3, detalle.cantidad)
                details_sheet.write_number(row_num, 4, detalle.costo_unitario, currency_format)
                details_sheet.write_number(row_num, 5, detalle.cantidad * detalle.costo_unitario, currency_format)
                row_num += 1
        
        # Ajustar anchos detalles
        details_sheet.set_column('A:A', 8)   # Compra ID
        details_sheet.set_column('B:B', 30)  # Producto
        details_sheet.set_column('C:C', 25)  # Proveedor
        details_sheet.set_column('D:D', 10)  # Cantidad
        details_sheet.set_column('E:F', 15)  # Precios
        
        workbook.close()
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=reporte_compras.xlsx'
        return response
        
    except Exception as e:
        app.logger.error(f'Error al generar Excel de compras: {str(e)}')
        flash('Error al generar el reporte de compras en Excel', 'danger')
        return redirect(url_for('admin_compras'))





@app.route('/admin/descargas/ventas_excel')
@admin_required
def descargar_ventas_excel():
    try:
        pedidos = Pedido.query.order_by(Pedido.fecha_pedido.desc()).all()
        
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet('Ventas')
        
        # Formatos
        header_format = workbook.add_format({
            'bold': True, 'bg_color': '#4472C4', 'font_color': 'white', 
            'border': 1, 'align': 'center'
        })
        date_format = workbook.add_format({'num_format': 'dd/mm/yyyy'})
        currency_format = workbook.add_format({'num_format': '$#,##0.00'})
        
        # Encabezados
        headers = [
            'ID', 'Fecha', 'Cliente', 'Total', 'Método Pago', 
            'Estado', 'Puntos Usados', 'Puntos Ganados'
        ]
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)
        
        # Datos
        for row_num, pedido in enumerate(pedidos, start=1):
            data = [
                pedido.id,
                pedido.fecha_pedido,
                pedido.usuario.nombre if pedido.usuario else 'N/A',
                pedido.total,
                pedido.metodo_pago.capitalize(),
                pedido.estado.capitalize(),
                pedido.puntos_usados,
                pedido.puntos_ganados
            ]
            
            for col_num, value in enumerate(data):
                if col_num == 1:  # Fecha
                    worksheet.write_datetime(row_num, col_num, pedido.fecha_pedido, date_format)
                elif col_num == 3:  # Total
                    worksheet.write_number(row_num, col_num, value, currency_format)
                else:
                    worksheet.write(row_num, col_num, value)
        
        # Ajustar anchos
        worksheet.set_column('A:A', 8)   # ID
        worksheet.set_column('B:B', 15)  # Fecha
        worksheet.set_column('C:C', 25)  # Cliente
        worksheet.set_column('D:D', 15)  # Total
        worksheet.set_column('E:E', 15)  # Método Pago
        worksheet.set_column('F:F', 15)  # Estado
        worksheet.set_column('G:H', 12)  # Puntos
        
        # Hoja de Detalles
        details_sheet = workbook.add_worksheet('Detalles')
        
        # Encabezados detalles
        detail_headers = [
            'Pedido ID', 'Producto', 'Cantidad', 'Precio Unitario',
            'Descuento', 'Subtotal'
        ]
        for col_num, header in enumerate(detail_headers):
            details_sheet.write(0, col_num, header, header_format)
        
        # Datos detalles
        row_num = 1
        for pedido in pedidos:
            for detalle in pedido.detalles:
                details_sheet.write(row_num, 0, pedido.id)
                details_sheet.write(row_num, 1, detalle.producto.nombre)
                details_sheet.write_number(row_num, 2, detalle.cantidad)
                details_sheet.write_number(row_num, 3, detalle.precio, currency_format)
                details_sheet.write_number(row_num, 4, detalle.descuento_aplicado/100 if detalle.descuento_aplicado > 0 else 0)
                details_sheet.write_number(row_num, 5, detalle.cantidad * detalle.precio, currency_format)
                row_num += 1
        
        # Ajustar anchos detalles
        details_sheet.set_column('A:A', 8)   # Pedido ID
        details_sheet.set_column('B:B', 30)  # Producto
        details_sheet.set_column('C:C', 10)  # Cantidad
        details_sheet.set_column('D:F', 15)  # Precios
        
        # Hoja de Estadísticas
        stats_sheet = workbook.add_worksheet('Estadísticas')
        
        total_ventas = sum(p.total for p in pedidos)
        ventas_completadas = sum(1 for p in pedidos if p.estado == 'completado')
        puntos_usados = sum(p.puntos_usados for p in pedidos)
        puntos_ganados = sum(p.puntos_ganados for p in pedidos)
        
        stats_data = [
            ['Total Ventas', len(pedidos)],
            ['Ventas Completadas', ventas_completadas],
            ['Total Recaudado', total_ventas],
            ['Puntos Usados', puntos_usados],
            ['Puntos Ganados', puntos_ganados]
        ]
        
        for row_num, (label, value) in enumerate(stats_data):
            stats_sheet.write(row_num, 0, label)
            if label == 'Total Recaudado':
                stats_sheet.write_number(row_num, 1, value, currency_format)
            else:
                stats_sheet.write(row_num, 1, value)
        
        workbook.close()
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=reporte_ventas.xlsx'
        return response
        
    except Exception as e:
        app.logger.error(f'Error al generar Excel de ventas: {str(e)}')
        flash('Error al generar el reporte de ventas en Excel', 'danger')
        return redirect(url_for('admin_pedidos'))



@app.route('/admin/descargas/proveedores_excel')
@admin_required
def descargar_proveedores_excel():
    try:
        proveedores = Proveedor.query.order_by(Proveedor.nombre).all()
        
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet('Proveedores')
        
        # Formatos
        header_format = workbook.add_format({
            'bold': True, 'bg_color': '#4472C4', 'font_color': 'white', 
            'border': 1, 'align': 'center'
        })
        
        # Encabezados
        headers = [
            'ID', 'Nombre', 'Identificación', 'Tipo', 'Email', 
            'Teléfono', 'Activo', 'Productos'
        ]
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)
        
        # Datos
        for row_num, proveedor in enumerate(proveedores, start=1):
            data = [
                proveedor.id,
                proveedor.nombre,
                proveedor.identificacion,
                proveedor.tipo,
                proveedor.email or 'N/A',
                proveedor.telefono or 'N/A',
                'Sí' if proveedor.activo else 'No',
                proveedor.productos.count()
            ]
            
            for col_num, value in enumerate(data):
                worksheet.write(row_num, col_num, value)
        
        # Ajustar anchos
        worksheet.set_column('A:A', 8)   # ID
        worksheet.set_column('B:B', 25)  # Nombre
        worksheet.set_column('C:C', 15)  # Identificación
        worksheet.set_column('D:D', 15)  # Tipo
        worksheet.set_column('E:E', 25)  # Email
        worksheet.set_column('F:F', 15)  # Teléfono
        worksheet.set_column('G:G', 10)  # Activo
        worksheet.set_column('H:H', 10)  # Productos
        
        # Hoja de Productos
        products_sheet = workbook.add_worksheet('Productos')
        
        # Encabezados productos
        product_headers = [
            'Proveedor', 'Producto', 'Marca', 'Costo', 'Stock'
        ]
        for col_num, header in enumerate(product_headers):
            products_sheet.write(0, col_num, header, header_format)
        
        # Datos productos
        row_num = 1
        for proveedor in proveedores:
            for producto in proveedor.productos:
                products_sheet.write(row_num, 0, proveedor.nombre)
                products_sheet.write(row_num, 1, producto.nombre)
                products_sheet.write(row_num, 2, producto.marca or 'N/A')
                products_sheet.write(row_num, 3, producto.costo)
                products_sheet.write(row_num, 4, producto.stock)
                row_num += 1
        
        # Ajustar anchos productos
        products_sheet.set_column('A:A', 25)  # Proveedor
        products_sheet.set_column('B:B', 30)  # Producto
        products_sheet.set_column('C:C', 20)  # Marca
        products_sheet.set_column('D:E', 15)  # Costo y Stock
        
        workbook.close()
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=reporte_proveedores.xlsx'
        return response
        
    except Exception as e:
        app.logger.error(f'Error al generar Excel de proveedores: {str(e)}')
        flash('Error al generar el reporte de proveedores en Excel', 'danger')
        return redirect(url_for('admin_proveedores'))

@app.route('/admin/descargas/usuarios_excel')
@admin_required
def descargar_usuarios_excel():
    try:
        usuarios = Usuario.query.order_by(Usuario.nombre).all()
        
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet('Usuarios')
        
        # Formatos
        header_format = workbook.add_format({
            'bold': True, 'bg_color': '#4472C4', 'font_color': 'white',
            'border': 1, 'align': 'center'
        })
        date_format = workbook.add_format({'num_format': 'dd/mm/yyyy hh:mm'})
        
        # Encabezados
        headers = [
            'ID', 'Nombre', 'Email', 'Identificación', 'Tipo',
            'Fecha Registro', 'Activo', 'Puntos'
        ]
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)
        
        # Datos
        for row_num, usuario in enumerate(usuarios, start=1):
            data = [
                usuario.id,
                usuario.nombre or 'N/A',
                usuario.email or 'N/A',
                usuario.identificacion or 'N/A',
                'Administrador' if usuario.es_admin else 'Cliente',
                usuario.fecha_registro,
                'Sí' if usuario.activo else 'No',
                usuario.puntos or 0
            ]
            
            for col_num, value in enumerate(data):
                if col_num == 5:  # Fecha Registro
                    if value:  # Verificar que la fecha no sea None
                        worksheet.write_datetime(row_num, col_num, value, date_format)
                    else:
                        worksheet.write(row_num, col_num, 'N/A')
                else:
                    worksheet.write(row_num, col_num, value)
        
        # Ajustar anchos
        worksheet.set_column('A:A', 8)   # ID
        worksheet.set_column('B:B', 25)  # Nombre
        worksheet.set_column('C:C', 30)  # Email
        worksheet.set_column('D:D', 15)  # Identificación
        worksheet.set_column('E:E', 15)  # Tipo
        worksheet.set_column('F:F', 20)  # Fecha Registro
        worksheet.set_column('G:G', 10)  # Activo
        worksheet.set_column('H:H', 10)  # Puntos
        
        workbook.close()
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=reporte_usuarios_{datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        app.logger.error(f'Error al generar Excel de usuarios: {str(e)}', exc_info=True)
        flash(f'Error al generar el reporte de usuarios en Excel: {str(e)}', 'danger')
        return redirect(url_for('admin_usuarios'))



@app.route('/admin/inventario/descargar_excel')
@admin_required
def descargar_reporte_inventario_excel():
    try:
        # Obtener parámetros de filtrado si los hay
        categoria_id = request.args.get('categoria_id', type=int)
        stock_filter = request.args.get('stock_filter', 'all')
        
        # Construir consulta base
        query = Producto.query.filter_by(activo=True)
        
        # Aplicar filtros
        if categoria_id:
            categoria = Categoria.query.get(categoria_id)
            if categoria:
                if categoria.parent_id is None:  # Es categoría principal
                    subcategorias_ids = [c.id for c in categoria.subcategorias]
                    query = query.filter(Producto.categoria_id.in_(subcategorias_ids))
                else:  # Es subcategoría
                    query = query.filter_by(categoria_id=categoria_id)
        
        if stock_filter == 'low':
            query = query.filter(Producto.stock < Producto.stock_minimo)
        elif stock_filter == 'over':
            query = query.filter(Producto.stock > Producto.stock_maximo)
        elif stock_filter == 'zero':
            query = query.filter(Producto.stock == 0)
        
        productos = query.order_by(Producto.nombre).all()
        
        # Crear libro de Excel
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet('Inventario')
        
        # Formatos
        header_format = workbook.add_format({
            'bold': True,
            'align': 'center',
            'valign': 'vcenter',
            'bg_color': '#4472C4',
            'font_color': 'white',
            'border': 1
        })
        
        currency_format = workbook.add_format({'num_format': '$#,##0.00'})
        percent_format = workbook.add_format({'num_format': '0.00%'})
        integer_format = workbook.add_format({'num_format': '0'})
        
        # Encabezados
        headers = [
            'ID', 'Nombre', 'Marca', 'Categoría', 'Subcategoría',
            'Stock', 'Stock Mínimo', 'Stock Máximo',
            'Precio Base', 'Precio Final', 'Descuento',
            'Valor Total'
        ]
        
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)
        
        # Datos
        for row_num, producto in enumerate(productos, start=1):
            categoria_principal = producto.categoria.parent.nombre if producto.categoria.parent else producto.categoria.nombre
            subcategoria = producto.categoria.nombre if producto.categoria.parent else '-'
            
            data = [
                producto.id,
                producto.nombre,
                producto.marca or '-',
                categoria_principal,
                subcategoria,
                producto.stock,
                producto.stock_minimo,
                producto.stock_maximo,
                producto.precio,
                producto.precio * (1 - producto.descuento/100) if producto.descuento > 0 else producto.precio,
                producto.descuento/100 if producto.descuento > 0 else 0,
                producto.stock * (producto.precio * (1 - producto.descuento/100) if producto.descuento > 0 else producto.precio)
            ]
            
            for col_num, value in enumerate(data):
                if col_num in [5, 6, 7]:  # Stock columns
                    worksheet.write_number(row_num, col_num, value, integer_format)
                elif col_num in [8, 9, 11]:  # Currency columns
                    worksheet.write_number(row_num, col_num, value, currency_format)
                elif col_num == 10:  # Discount column
                    worksheet.write_number(row_num, col_num, value, percent_format)
                else:
                    worksheet.write(row_num, col_num, value)
        
        # Ajustar anchos de columna
        worksheet.set_column('A:A', 8)   # ID
        worksheet.set_column('B:B', 30)  # Nombre
        worksheet.set_column('C:C', 20)  # Marca
        worksheet.set_column('D:D', 20)  # Categoría
        worksheet.set_column('E:E', 20)  # Subcategoría
        worksheet.set_column('F:H', 12)  # Stocks
        worksheet.set_column('I:J', 15)  # Precios
        worksheet.set_column('K:K', 12)  # Descuento
        worksheet.set_column('L:L', 15)  # Valor Total
        
        # Hoja de Resumen
        summary_sheet = workbook.add_worksheet('Resumen')
        
        # Estadísticas
        total_productos = len(productos)
        productos_bajo_stock = len([p for p in productos if p.stock < p.stock_minimo])
        productos_sobre_stock = len([p for p in productos if p.stock > p.stock_maximo])
        valor_total = sum(p.precio * p.stock * (1 - p.descuento/100) for p in productos)
        
        summary_data = [
            ['Total Productos', total_productos],
            ['Productos bajo stock mínimo', productos_bajo_stock],
            ['Productos sobre stock máximo', productos_sobre_stock],
            ['Valor total del inventario', valor_total]
        ]
        
        for row_num, (label, value) in enumerate(summary_data):
            summary_sheet.write(row_num, 0, label)
            if isinstance(value, (int, float)):
                if label == 'Valor total del inventario':
                    summary_sheet.write_number(row_num, 1, value, currency_format)
                else:
                    summary_sheet.write_number(row_num, 1, value)
            else:
                summary_sheet.write(row_num, 1, value)
        
        # Gráfico de distribución por categoría
        chart_sheet = workbook.add_worksheet('Gráficos')
        
        # Datos para gráfico
        categorias = defaultdict(float)
        for p in productos:
            cat_nombre = p.categoria.parent.nombre if p.categoria.parent else p.categoria.nombre
            valor = p.precio * p.stock * (1 - p.descuento/100)
            categorias[cat_nombre] += valor
        
        # Escribir datos para el gráfico
        chart_sheet.write_column('A1', list(categorias.keys()))
        chart_sheet.write_column('B1', list(categorias.values()), currency_format)
        
        # Crear gráfico
        chart = workbook.add_chart({'type': 'pie'})
        chart.add_series({
            'categories': '=Gráficos!$A$1:$A${}'.format(len(categorias)),
            'values': '=Gráficos!$B$1:$B${}'.format(len(categorias)),
            'name': 'Distribución por categoría'
        })
        chart.set_title({'name': 'Valor del Inventario por Categoría'})
        chart_sheet.insert_chart('D2', chart)
        
        workbook.close()
        
        # Crear respuesta
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        filename = 'reporte_inventario_{}.xlsx'.format(datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S'))
        response.headers['Content-Disposition'] = 'attachment; filename={}'.format(filename)
        
        return response
    
    except Exception as e:
        app.logger.error(f'Error al generar reporte Excel: {str(e)}')
        flash('Error al generar el reporte en Excel', 'danger')
        return redirect(url_for('reportes_inventario'))

@app.template_filter('currency')
def currency_filter(value):
    try:
        value = float(value)
        return 'COP ' + '{:,.2f}'.format(value).replace(',', 'X').replace('.', ',').replace('X', '.')
    except Exception:
        return value

@app.route('/admin/productos/sugerencias')
@admin_required
def sugerencias_productos_admin():
    campo = request.args.get('campo', 'nombre')
    query = request.args.get('q', '').strip()
    results = []
    if query and campo in ['nombre', 'descripcion']:
        if campo == 'nombre':
            productos = Producto.query.filter(Producto.nombre.ilike(f'%{query}%')).limit(8).all()
            results = [{'id': p.id, 'text': p.nombre} for p in productos]
        elif campo == 'descripcion':
            productos = Producto.query.filter(Producto.descripcion.ilike(f'%{query}%')).limit(8).all()
            results = [{'id': p.id, 'text': p.descripcion} for p in productos if p.descripcion]
    return jsonify(results)

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        if not Categoria.query.first():
            admin = Usuario(
                nombre="Admin",
                email="admin@tienda.com",
                identificacion="1041771628",
                password=generate_password_hash("admin123"),
                es_admin=True,
                fecha_registro=datetime.now(timezone.utc)
            )
            db.session.add(admin)
            db.session.commit()

    app.run(debug=True)